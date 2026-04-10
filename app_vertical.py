# app_vertical.py

import json
import os
import threading
import time
from decimal import Decimal, ROUND_HALF_UP
from functools import partial

import numpy as np
import pandas as pd
import tkinter as tk
from openpyxl.styles import Alignment, Border, Side
from tkinter import filedialog, messagebox, simpledialog, ttk
from tkinterdnd2 import DND_FILES

from common_utils import (
    EXTENDED_STATS,
    FILTER_OPERATORS,
    apply_filter_conditions,
    apply_count_masks,
    build_grouped_stats_frame,
    calculate_series_stats,
    describe_filter_conditions,
    open_with_default_app,
    read_csv_safely,
    round2,
    sort_key,
)
from ui_shell import (
    begin_busy,
    bind_shortcuts,
    build_output_path,
    build_app_menu,
    copy_text,
    end_busy,
    initialize_shell,
    mark_output,
    maybe_restore_recent_file,
    open_output_folder,
    populate_recent_menus,
    set_status,
    show_workbook_preview,
)

# 当前程序版本号——每次发布时请手动更新
APP_VERSION = "1.2.0"

UPDATE_CONTENT = """
小捞翔·至尊版 v1.2.0 更新日志：
- 新增：扩展统计指标，支持缺失数、缺失率、四分位数、极差、方差、偏度、峰度等
- 新增：自动生成“统计说明”和“导出清单”工作表，方便回看参数和输出内容
- 新增：值字段会自动尝试转为数值，无法计算的组合会提示并自动跳过
- 优化：横版和竖版统计口径统一，更多统计方法在两个界面保持一致
- 优化：统计量选择面板扩大，容纳更多指标
- 优化：取消批量选项后会自动清空对应字段，减少误操作
"""

# 默认配置文件名（可自定义目录）
CONFIG_FILE = "small_shit_vertical.json"

MAX_LEVELS = 3  # 最多三级分类




class VerticalApp:
    def __init__(self, root):
        self.root = root
        initialize_shell(self, mode_name="vertical", title="小捞翔")
        # -------- 新增：自定义排序状态 --------
        self.custom_orders = {}  # 用户自定义的每列排序列表
        self.original_orders = {}  # 首次记录每列原始唯一值顺序
        # -------- 新增 recent_files 存储 最近打开 列表 --------
        self.recent_files = []
        root.geometry("720x720")
        root.minsize(620, 640)
        root.resizable(True, True)

        # 在窗口关闭前保存配置
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        root.drop_target_register(DND_FILES)
        root.dnd_bind("<<DragEnter>>", lambda e: root.configure(bg="#1f232a"))
        root.dnd_bind("<<DragLeave>>", lambda e: root.configure(bg="SystemButtonFace"))
        root.dnd_bind("<<Drop>>", self.handle_drop)

        # 初始化属性
        self.excel_path = ""
        self.data = pd.DataFrame()
        self.batch_fields = []
        self.group_batch_fields = []
        self.filter_conditions = []
        self.group_templates = {}
        self.active_group_template_name = ""
        self.update_history = []  # 更新历史记录
        self.shown_version = ""  # 已显示的版本

        # 保存上次选择的统计量，初始全选
        self.all_stats = EXTENDED_STATS.copy()
        self.last_stats = self.all_stats.copy()

        # 构建菜单栏
        self._build_menu()
        self._build_file_sheet_selector()
        self._build_levels_frame()
        self._build_field_controls()
        self._build_batch_controls()  # 添加批量控制组件
        self._build_buttons()
        self._build_status_bar()

        # 载入上次保存的配置
        self.load_config()
        self.create_recent_menu()
        self.toggle_area_fields()
        bind_shortcuts(self)
        # 检查更新
        self.check_for_update()
        maybe_restore_recent_file(self)

    def on_closing(self):
        # 退出前保存配置
        self.save_config(show_msg=False)
        self.root.destroy()

    def _normalize_filter_conditions(self, cols):
        normalized = []
        for condition in getattr(self, "filter_conditions", []):
            if not isinstance(condition, dict):
                continue
            field = str(condition.get("field", "")).strip()
            op = str(condition.get("op", "")).strip()
            if field and op and field in cols:
                normalized.append(
                    {
                        "field": field,
                        "op": op,
                        "value": str(condition.get("value", "")),
                        "enabled": bool(condition.get("enabled", True)),
                    }
                )
        self.filter_conditions = normalized

    def _filter_summary(self):
        active_conditions = [c for c in self.filter_conditions if c.get("enabled", True)]
        return describe_filter_conditions(active_conditions)

    def _get_filtered_data(self, show_error=True):
        if self.data.empty:
            return self.data.copy()

        try:
            return apply_filter_conditions(self.data, self.filter_conditions)
        except Exception as exc:
            if show_error:
                messagebox.showerror("筛选条件有误", str(exc))
                set_status(self, "筛选条件有误，请检查后再试")
            return None

    def get_active_data(self):
        filtered = self._get_filtered_data(show_error=False)
        if isinstance(filtered, pd.DataFrame):
            return filtered
        return self.data

    def _update_loaded_status(self):
        if self.data.empty:
            set_status(self, f"{self.mode_label} 就绪")
            return

        total_rows = len(self.data)
        total_cols = len(self.data.columns)
        if not self.filter_conditions:
            set_status(self, f"已加载 {total_rows} 行 / {total_cols} 列")
            return

        filtered = self._get_filtered_data(show_error=False)
        if filtered is None:
            set_status(self, f"已加载 {total_rows} 行 / {total_cols} 列，筛选条件待修正")
            return

        set_status(self, f"已加载 {total_rows} 行 / {total_cols} 列，筛选后 {len(filtered)} 行")

    def _collect_group_template(self):
        return {
            "ui_state": self._collect_ui_state(),
            "stats": list(self.last_stats),
        }

    def _describe_group_template(self, name, payload):
        state = payload.get("ui_state", {}) if isinstance(payload, dict) else {}
        stats = payload.get("stats", []) if isinstance(payload, dict) else []
        lines = [
            f"模板名称：{name}",
            f"分类级别：{' / '.join(state.get('levels', [])) or '未设置'}",
            f"值字段：{state.get('value_field') or '未设置'}",
            f"面积字段：{state.get('ratio_field') or '未设置'}",
            f"面积统计：{'是' if state.get('area_enabled', False) else '否'}",
            f"批量值字段：{' / '.join(state.get('batch_fields', [])) or '未启用'}",
            f"批量分组字段：{' / '.join(state.get('group_batch_fields', [])) or '未启用'}",
            f"统计量：{' / '.join(stats) if stats else '未设置'}",
        ]
        return "\n".join(lines)

    def _apply_group_template(self, payload, template_name=""):
        if self.data.empty:
            messagebox.showwarning("提示", "请先加载数据后再应用模板")
            return False

        if not isinstance(payload, dict):
            messagebox.showerror("错误", "模板内容无效")
            return False

        state = payload.get("ui_state", {})
        if not isinstance(state, dict):
            messagebox.showerror("错误", "模板缺少界面配置")
            return False

        cols = list(self.data.columns)
        enabled_fields = list(state.get("levels", []))
        if state.get("value_field"):
            enabled_fields.append(state.get("value_field"))
        if state.get("area_enabled", False) and state.get("ratio_field"):
            enabled_fields.append(state.get("ratio_field"))
        if state.get("batch_enabled", False):
            enabled_fields.extend(state.get("batch_fields", []))
        if state.get("group_batch_enabled", False):
            enabled_fields.extend(state.get("group_batch_fields", []))

        missing_fields = sorted({field for field in enabled_fields if field and field not in cols})
        if missing_fields:
            messagebox.showwarning("模板无法应用", "当前数据缺少以下字段：\n" + "\n".join(missing_fields))
            return False

        self.saved_ui_state = state
        self._restore_ui_state(cols)
        stats = [stat for stat in payload.get("stats", []) if stat in self.all_stats]
        if stats:
            self.last_stats = stats
        self.active_group_template_name = template_name
        self._update_loaded_status()
        return True

    def open_group_template_manager(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("分组模板")
        dialog.geometry("720x420")
        dialog.transient(self.root)

        container = ttk.Frame(dialog, padding=12)
        container.pack(fill="both", expand=True)
        container.columnconfigure(1, weight=1)
        container.rowconfigure(0, weight=1)

        left = ttk.Frame(container)
        left.grid(row=0, column=0, sticky="ns", padx=(0, 12))
        right = ttk.Frame(container)
        right.grid(row=0, column=1, sticky="nsew")
        right.rowconfigure(1, weight=1)

        ttk.Label(left, text="已保存模板").pack(anchor="w")
        listbox = tk.Listbox(left, width=24, height=16)
        listbox.pack(fill="y", expand=True, pady=(6, 8))

        ttk.Label(right, text="模板内容").grid(row=0, column=0, sticky="w")
        preview = tk.Text(right, wrap="word", state="disabled", height=16)
        preview.grid(row=1, column=0, sticky="nsew", pady=(6, 8))

        buttons = ttk.Frame(right)
        buttons.grid(row=2, column=0, sticky="e")

        def selected_name():
            selection = listbox.curselection()
            if not selection:
                return ""
            return listbox.get(selection[0])

        def refresh_templates(target_name=""):
            names = sorted(self.group_templates)
            listbox.delete(0, "end")
            for name in names:
                listbox.insert("end", name)

            if target_name and target_name in names:
                index = names.index(target_name)
                listbox.selection_set(index)
                listbox.see(index)
            elif names:
                listbox.selection_set(0)

            update_preview()

        def update_preview(_event=None):
            name = selected_name()
            content = "暂无模板"
            if name:
                content = self._describe_group_template(name, self.group_templates.get(name, {}))
            preview.configure(state="normal")
            preview.delete("1.0", "end")
            preview.insert("1.0", content)
            preview.configure(state="disabled")

        def save_current_template():
            if self.data.empty:
                messagebox.showwarning("提示", "请先加载数据后再保存模板")
                return

            initial_name = self.active_group_template_name or selected_name() or "常用模板"
            name = simpledialog.askstring("保存模板", "请输入模板名称：", parent=dialog, initialvalue=initial_name)
            if not name:
                return
            name = name.strip()
            if not name:
                return

            self.group_templates[name] = self._collect_group_template()
            self.active_group_template_name = name
            self.save_config(show_msg=False)
            refresh_templates(name)
            set_status(self, f"模板“{name}”已保存")

        def apply_selected_template():
            name = selected_name()
            if not name:
                messagebox.showinfo("提示", "请先选择一个模板")
                return
            if self._apply_group_template(self.group_templates.get(name, {}), name):
                self.save_config(show_msg=False)
                dialog.destroy()

        def delete_selected_template():
            name = selected_name()
            if not name:
                return
            if not messagebox.askyesno("确认删除", f"确定要删除模板“{name}”吗？", parent=dialog):
                return
            self.group_templates.pop(name, None)
            if self.active_group_template_name == name:
                self.active_group_template_name = ""
            self.save_config(show_msg=False)
            refresh_templates()

        def export_selected_template():
            name = selected_name()
            if not name:
                messagebox.showinfo("提示", "请先选择一个模板")
                return
            path = filedialog.asksaveasfilename(
                title="导出模板",
                defaultextension=".json",
                filetypes=[("JSON 文件", "*.json"), ("所有文件", "*.*")],
                initialfile=f"{name}.json",
            )
            if not path:
                return
            with open(path, "w", encoding="utf-8") as file:
                json.dump({"name": name, "template": self.group_templates[name]}, file, ensure_ascii=False, indent=2)
            set_status(self, f"模板已导出到 {os.path.basename(path)}")

        def import_template():
            path = filedialog.askopenfilename(
                title="导入模板",
                filetypes=[("JSON 文件", "*.json"), ("所有文件", "*.*")],
            )
            if not path:
                return

            with open(path, "r", encoding="utf-8") as file:
                payload = json.load(file)

            if isinstance(payload, dict) and "template" in payload:
                name = str(payload.get("name") or os.path.splitext(os.path.basename(path))[0]).strip()
                template = payload.get("template", {})
            else:
                name = os.path.splitext(os.path.basename(path))[0]
                template = payload

            if not name or not isinstance(template, dict):
                messagebox.showerror("导入失败", "模板文件格式不正确")
                return

            self.group_templates[name] = template
            self.save_config(show_msg=False)
            refresh_templates(name)
            set_status(self, f"模板“{name}”已导入")

        ttk.Button(buttons, text="保存当前", command=save_current_template).pack(side="left", padx=(0, 6))
        ttk.Button(buttons, text="应用", command=apply_selected_template).pack(side="left", padx=6)
        ttk.Button(buttons, text="删除", command=delete_selected_template).pack(side="left", padx=6)
        ttk.Button(buttons, text="导出", command=export_selected_template).pack(side="left", padx=6)
        ttk.Button(buttons, text="导入", command=import_template).pack(side="left", padx=6)
        ttk.Button(buttons, text="关闭", command=dialog.destroy).pack(side="left", padx=(6, 0))

        listbox.bind("<<ListboxSelect>>", update_preview)
        refresh_templates(self.active_group_template_name)
        dialog.grab_set()
        self.root.wait_window(dialog)

    def open_filter_builder(self):
        if self.data.empty:
            messagebox.showwarning("提示", "请先加载数据后再设置筛选条件")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("条件筛选")
        dialog.geometry("760x420")
        dialog.transient(self.root)

        container = ttk.Frame(dialog, padding=12)
        container.pack(fill="both", expand=True)
        container.rowconfigure(1, weight=1)
        container.columnconfigure(0, weight=1)

        ttk.Label(container, text=f"当前数据共 {len(self.data)} 行，可添加多个筛选条件。").grid(row=0, column=0, sticky="w")

        rows_frame = ttk.Frame(container)
        rows_frame.grid(row=1, column=0, sticky="nsew", pady=(10, 8))
        rows_frame.columnconfigure(0, weight=1)

        summary_var = tk.StringVar()
        detail_var = tk.StringVar(value=self._filter_summary())
        condition_rows = []
        cols = list(self.data.columns)

        def collect_conditions():
            conditions = []
            for item in condition_rows:
                field = item["field"].get().strip()
                op = item["op"].get().strip()
                value = item["value"].get().strip()
                enabled = bool(item["enabled"].get())
                if field and op:
                    conditions.append(
                        {
                            "field": field,
                            "op": op,
                            "value": value,
                            "enabled": enabled,
                        }
                    )
            return conditions

        def refresh_summary():
            conditions = collect_conditions()
            active_conditions = [c for c in conditions if c.get("enabled", True)]
            detail_var.set(describe_filter_conditions(active_conditions))
            try:
                filtered = apply_filter_conditions(self.data, conditions)
            except Exception as exc:
                summary_var.set(f"条件有误：{exc}")
                return

            if active_conditions:
                summary_var.set(f"已启用 {len(active_conditions)} 条，筛选后 {len(filtered)} / {len(self.data)} 行")
            else:
                summary_var.set(f"当前未启用筛选，将使用全部 {len(self.data)} 行数据")

        def remove_condition(item):
            item["frame"].destroy()
            if item in condition_rows:
                condition_rows.remove(item)
            if not condition_rows:
                add_condition_row()
            refresh_summary()

        def add_condition_row(condition=None):
            condition = condition or {}
            frame = ttk.Frame(rows_frame)
            frame.pack(fill="x", pady=4)

            enabled_var = tk.BooleanVar(value=bool(condition.get("enabled", True)))
            field_var = tk.StringVar(value=str(condition.get("field", cols[0] if cols else "")))
            op_var = tk.StringVar(value=str(condition.get("op", FILTER_OPERATORS[0])))
            value_var = tk.StringVar(value=str(condition.get("value", "")))

            ttk.Checkbutton(frame, text="启用", variable=enabled_var, command=refresh_summary).pack(side="left")
            field_box = ttk.Combobox(frame, values=cols, textvariable=field_var, state="readonly", width=20)
            field_box.pack(side="left", padx=(8, 6))
            op_box = ttk.Combobox(frame, values=FILTER_OPERATORS, textvariable=op_var, state="readonly", width=12)
            op_box.pack(side="left", padx=6)
            value_entry = ttk.Entry(frame, textvariable=value_var, width=22)
            value_entry.pack(side="left", padx=6)

            item = {
                "frame": frame,
                "enabled": enabled_var,
                "field": field_var,
                "op": op_var,
                "value": value_var,
            }

            def sync_value_state(*_args):
                disabled = op_var.get() in {"为空", "不为空"}
                value_entry.configure(state="disabled" if disabled else "normal")
                if disabled:
                    value_var.set("")
                refresh_summary()

            op_var.trace_add("write", sync_value_state)
            field_box.bind("<<ComboboxSelected>>", lambda _event: refresh_summary())
            value_entry.bind("<KeyRelease>", lambda _event: refresh_summary())
            ttk.Button(frame, text="删除", command=lambda: remove_condition(item)).pack(side="left", padx=(6, 0))
            condition_rows.append(item)
            sync_value_state()

        controls = ttk.Frame(container)
        controls.grid(row=2, column=0, sticky="ew")
        ttk.Button(controls, text="新增条件", command=add_condition_row).pack(side="left")
        ttk.Button(
            controls,
            text="清空筛选",
            command=lambda: (
                [row["frame"].destroy() for row in condition_rows],
                condition_rows.clear(),
                self.filter_conditions.clear(),
                add_condition_row(),
                refresh_summary(),
            ),
        ).pack(side="left", padx=6)

        ttk.Label(container, textvariable=summary_var).grid(row=3, column=0, sticky="w", pady=(10, 2))
        ttk.Label(container, textvariable=detail_var, wraplength=700).grid(row=4, column=0, sticky="w")

        buttons = ttk.Frame(container)
        buttons.grid(row=5, column=0, sticky="e", pady=(12, 0))

        def apply_conditions():
            conditions = collect_conditions()
            try:
                filtered = apply_filter_conditions(self.data, conditions)
            except Exception as exc:
                messagebox.showerror("筛选条件有误", str(exc), parent=dialog)
                return

            self.filter_conditions = [c for c in conditions if c.get("enabled", True)]
            self._update_loaded_status()
            self.save_config(show_msg=False)
            if self.filter_conditions:
                set_status(self, f"筛选已更新，当前保留 {len(filtered)} 行")
            else:
                set_status(self, f"已清除筛选，当前共 {len(self.data)} 行")
            dialog.destroy()

        ttk.Button(buttons, text="应用", command=apply_conditions).pack(side="left", padx=(0, 6))
        ttk.Button(buttons, text="关闭", command=dialog.destroy).pack(side="left")

        existing_conditions = self.filter_conditions or [{}]
        for condition in existing_conditions:
            add_condition_row(condition)
        refresh_summary()
        dialog.grab_set()
        self.root.wait_window(dialog)

    def _collect_ui_state(self):
        return {
            "levels": [lvl["var"].get() for lvl in self.levels if lvl["var"].get()],
            "value_field": self.val_var.get(),
            "ratio_field": self.ratio_var.get(),
            "area_enabled": bool(self.area_cb.get()),
            "batch_enabled": bool(self.batch_var.get()),
            "batch_fields": list(self.batch_fields),
            "group_batch_enabled": bool(self.group_batch_var.get()),
            "group_batch_fields": list(self.group_batch_fields),
        }

    def _set_level_count(self, target_count):
        target_count = max(1, min(MAX_LEVELS, target_count))
        while len(self.levels) < target_count:
            self.add_level()
        while len(self.levels) > target_count:
            self._remove_level(self.levels[-1]["btn"])

    def _restore_ui_state(self, cols):
        state = getattr(self, "saved_ui_state", {})
        if not isinstance(state, dict):
            return

        saved_levels = [field for field in state.get("levels", []) if field in cols]
        if saved_levels:
            self._set_level_count(len(saved_levels))
            for idx, lvl in enumerate(self.levels):
                if idx < len(saved_levels):
                    lvl["var"].set(saved_levels[idx])
                elif cols:
                    lvl["var"].set(cols[0])

        saved_value = state.get("value_field")
        saved_ratio = state.get("ratio_field")
        if saved_value in cols:
            self.val_var.set(saved_value)
        if saved_ratio in cols:
            self.ratio_var.set(saved_ratio)

        self.area_cb.set(bool(state.get("area_enabled", self.area_cb.get())))
        self.toggle_area_fields()
        if self.area_cb.get() and saved_ratio in cols:
            self.ratio_var.set(saved_ratio)

        self.batch_var.set(bool(state.get("batch_enabled", False)))
        self.toggle_batch()
        self.batch_fields = [field for field in state.get("batch_fields", []) if field in cols]

        self.group_batch_var.set(bool(state.get("group_batch_enabled", False)))
        self.toggle_group_batch()
        self.group_batch_fields = [field for field in state.get("group_batch_fields", []) if field in cols]

    def save_config(self, path=None, show_msg=True):
        """保存配置到JSON文件"""
        cfg = {
            "shown_version": getattr(self, "shown_version", ""),
            "update_history": getattr(self, "update_history", []),
            "last_stats": self.last_stats,
            "recent_files": self.recent_files
            , "custom_orders": self.custom_orders,
            "area_decimals": self.area_decimals_var.get(),
            "ui_state": self._collect_ui_state(),
            "filter_conditions": self.filter_conditions,
            "group_templates": self.group_templates,
            "active_group_template_name": self.active_group_template_name,
            "last_output_path": self.last_output_path,
        }
        cfg_path = path or CONFIG_FILE
        try:
            with open(cfg_path, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
            if show_msg:
                messagebox.showinfo("配置已保存", f"配置已保存到：\n{cfg_path}")
        except Exception as e:
            if show_msg:
                messagebox.showerror("保存失败", str(e))

    def load_config(self, path=None):
        """从JSON文件加载配置"""
        cfg_path = path or CONFIG_FILE
        if not os.path.exists(cfg_path):
            return

        try:
            with open(cfg_path, "r", encoding="utf-8") as f:
                cfg = json.load(f)

            # 已展示过更新提示的版本号
            self.shown_version = cfg.get("shown_version", "")

            # 已累计的更新日志历史（多条）
            self.update_history = cfg.get("update_history", [])

            # 上次选择的统计量列表
            self.last_stats = cfg.get("last_stats", self.all_stats.copy())
            # 恢复最近打开列表 & 自定义排序
            self.recent_files = cfg.get("recent_files", [])
            self.custom_orders = cfg.get("custom_orders", {})
            dec = cfg.get("area_decimals", 2)
            self.area_decimals_var.set(dec)
            self.saved_ui_state = cfg.get("ui_state", {})
            self.filter_conditions = cfg.get("filter_conditions", [])
            self.group_templates = cfg.get("group_templates", {})
            self.active_group_template_name = cfg.get("active_group_template_name", "")
            self.last_output_path = cfg.get("last_output_path")
        # 配置载入后刷新最近打开菜单（如果已经创建）
            if hasattr(self, "create_recent_menu"):
                self.create_recent_menu()


        except Exception as e:
            messagebox.showerror("加载失败", str(e))

    def check_for_update(self):
        """检查并显示更新日志"""
        if self.shown_version != APP_VERSION:
            dlg = tk.Toplevel(self.root)
            dlg.title(f"更新日志 v{APP_VERSION}")
            dlg.geometry("520x360")
            dlg.resizable(False, False)

            # — 文本区 —
            frm_txt = ttk.Frame(dlg, padding=8)
            frm_txt.pack(fill="both", expand=True)
            txt = tk.Text(frm_txt, wrap="word", height=10)
            txt.insert("1.0", UPDATE_CONTENT.strip())
            txt.configure(state="disabled")  # 只读
            txt.pack(fill="both", expand=True)

            # — 按钮区 —
            frm_btn = ttk.Frame(dlg, padding=(8, 8))
            frm_btn.pack(fill="x")

            def _on_close():
                self.shown_version = APP_VERSION
                # 把本次更新内容加入历史（避免重复）
                if UPDATE_CONTENT.strip() not in self.update_history:
                    self.update_history.append(UPDATE_CONTENT.strip())
                self.save_config(show_msg=False)
                dlg.destroy()

            btn = ttk.Button(frm_btn, text="知道了", command=_on_close)
            btn.pack(side="right", padx=4, pady=4)

            dlg.transient(self.root)
            dlg.grab_set()
            self.root.wait_window(dlg)

    def show_update_history(self):
        """显示更新历史记录"""
        history = getattr(self, "update_history", [])
        if not history:
            messagebox.showinfo("更新历史", "暂无更新记录。")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("更新历史")
        dlg.geometry("520x360")
        dlg.resizable(False, False)

        frm = ttk.Frame(dlg, padding=8)
        frm.pack(fill="both", expand=True)

        txt = tk.Text(frm, wrap="word")
        txt.insert("1.0", "\n\n".join(history))
        txt.configure(state="disabled")
        txt.pack(fill="both", expand=True)

        btn = ttk.Button(frm, text="关闭", command=dlg.destroy)
        btn.pack(pady=6)

        dlg.transient(self.root)
        dlg.grab_set()
        self.root.wait_window(dlg)

    def _build_menu(self):
        build_app_menu(self)

    def create_recent_menu(self):
        """刷新 最近打开 下拉列表"""
        populate_recent_menus(self)

    def add_recent(self, path):
        """把 path 插入到 最近打开（去重 + 限长 20）"""
        if path in self.recent_files:
            self.recent_files.remove(path)
        self.recent_files.insert(0, path)
        self.recent_files = self.recent_files[:20]
        # 静默保存并刷新菜单
        self.save_config(show_msg=False)
        self.create_recent_menu()

    def open_recent(self, path):
        """通过 最近打开 重新载入文件"""
        if not os.path.exists(path):
            messagebox.showerror("文件不存在", path)
            self.recent_files.remove(path)
            self.save_config(show_msg=False)
            self.create_recent_menu()
            return

        self._load_path(path)

    def _build_file_sheet_selector(self):
        frm = ttk.Frame(self.root)
        frm.pack(fill="x", padx=10, pady=8)

        ttk.Button(frm, text="选择 Excel", command=self.select_file) \
            .grid(row=0, column=0, sticky="w")

        self.recent_btn = ttk.Menubutton(frm, text="最近打开")
        self.recent_menu = tk.Menu(self.recent_btn, tearoff=0)
        self.recent_btn["menu"] = self.recent_menu
        self.recent_btn.grid(row=0, column=1, padx=(8, 0), sticky="w")

        self.lbl_file = ttk.Label(frm, text="未选择文件")
        self.lbl_file.grid(row=0, column=2, padx=8, sticky="w")

        ttk.Label(frm, text="子表:") \
            .grid(row=1, column=0, pady=6, sticky="w")

        self.sheet_var = tk.StringVar()
        self.sheet_menu = ttk.Combobox(
            frm, textvariable=self.sheet_var,
            state="readonly", width=28
        )
        self.sheet_menu.grid(row=1, column=2, sticky="w")
        self.sheet_menu.bind(
            "<<ComboboxSelected>>",
            lambda e: self.load_sheet(self.sheet_var.get())
        )

    def _build_levels_frame(self):
        self.frame_levels = ttk.LabelFrame(
            self.root, text="分类级别 (1～3 级)"
        )
        self.frame_levels.pack(fill="x", padx=10, pady=6)

        self.levels = []
        self.add_btn = ttk.Button(
            self.frame_levels, text="+ 添加级别",
            command=self.add_level
        )
        self.add_level()
        self.add_level()

    def _build_field_controls(self):
        frm = ttk.Frame(self.root)
        frm.pack(fill="x", padx=10, pady=6)

        # 值字段
        ttk.Label(frm, text="值字段:") \
            .grid(row=0, column=0, sticky="w")
        self.val_var = tk.StringVar()
        self.val_menu = ttk.Combobox(
            frm, textvariable=self.val_var,
            state="readonly", width=28
        )
        self.val_menu.grid(row=0, column=1, sticky="w")

        # 面积占比字段
        ttk.Label(frm, text="面积占比字段:") \
            .grid(row=1, column=0, sticky="w")
        self.ratio_var = tk.StringVar()
        self.ratio_menu = ttk.Combobox(
            frm, textvariable=self.ratio_var,
            state="readonly", width=28
        )
        self.ratio_menu.grid(row=1, column=1, sticky="w")
        # 面积小数位
        ttk.Label(frm, text="面积小数位:") \
            .grid(row=2, column=0, sticky="w")

        self.area_decimals_var = tk.IntVar(value=2)
        self.area_decimals_spin = ttk.Spinbox(
            frm, from_=0, to=10, textvariable=self.area_decimals_var,
            width=5, state="readonly"
        )
        self.area_decimals_spin.grid(row=2, column=1, sticky="w")
        self.area_cb = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            frm, text="计算面积占比", variable=self.area_cb, command=self.toggle_area_fields
        ).grid(row=1, column=2, padx=5)

    def _build_batch_controls(self):
        """添加批量功能控件"""
        frm = ttk.LabelFrame(self.root, text="批量功能")
        frm.pack(fill="x", padx=10, pady=6)

        # 批量值字段
        self.batch_var = tk.BooleanVar(value=False)
        batch_frame = ttk.Frame(frm)
        batch_frame.pack(fill="x", padx=5, pady=5)
        ttk.Checkbutton(
            batch_frame, text="批量值字段",
            variable=self.batch_var,
            command=self.toggle_batch
        ).pack(side="left", padx=5)
        self.batch_btn = ttk.Button(
            batch_frame, text="选择值字段",
            command=self.choose_batch_fields,
            state="disabled"
        )
        self.batch_btn.pack(side="left", padx=5)

        # 批量分组字段
        self.group_batch_var = tk.BooleanVar(value=False)
        group_frame = ttk.Frame(frm)
        group_frame.pack(fill="x", padx=5, pady=5)
        # -------- 新增：自定义统计量顺序 按钮 --------
        btn = ttk.Button(frm, text="自定义统计量顺序", command=self.open_stats_order)
        btn.pack(side="right", padx=5)

        ttk.Checkbutton(
            group_frame, text="批量分组字段",
            variable=self.group_batch_var,
            command=self.toggle_group_batch
        ).pack(side="left", padx=5)
        self.group_batch_btn = ttk.Button(
            group_frame, text="选择分组字段",
            command=self.choose_group_batch_fields,
            state="disabled"
        )
        # -------- 新增：选择统计量 按钮，与横版保持一致 --------
        stats_btn = ttk.Button(frm, text="选择统计量", command=self.choose_stats)
        stats_btn.pack(side="right", padx=5)
        sort_btn = ttk.Button(frm, text="自定义排序", command=self.open_custom_sort)
        sort_btn.pack(side="right", padx=5)
        filter_btn = ttk.Button(frm, text="条件筛选", command=self.open_filter_builder)
        filter_btn.pack(side="right", padx=5)
        template_btn = ttk.Button(frm, text="分组模板", command=self.open_group_template_manager)
        template_btn.pack(side="right", padx=5)
        self.group_batch_btn.pack(side="left", padx=5)




    def toggle_batch(self):
        """切换批量值字段按钮状态"""
        if self.batch_var.get():
            self.batch_btn.state(["!disabled"])
        else:
            self.batch_btn.state(["disabled"])
            self.batch_fields = []

    def toggle_group_batch(self):
        """切换批量分组字段按钮状态"""
        if self.group_batch_var.get():
            self.group_batch_btn.state(["!disabled"])
        else:
            self.group_batch_btn.state(["disabled"])
            self.group_batch_fields = []

    def choose_batch_fields(self):
        """选择多个值字段"""
        if self.data.empty:
            messagebox.showwarning("提示", "请先加载数据")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("选择值字段(多选)")
        cols = list(self.data.columns)
        lb = tk.Listbox(dlg, selectmode="multiple", height=10)
        for c in cols:
            lb.insert("end", c)
        lb.pack(padx=12, pady=6)

        def on_ok():
            self.batch_fields = [cols[i] for i in lb.curselection()]
            dlg.destroy()

        ttk.Button(dlg, text="确定", command=on_ok, width=10).pack(pady=6)
        dlg.grab_set()
        self.root.wait_window(dlg)

    def choose_group_batch_fields(self):
        """选择多个分组字段"""
        if self.data.empty:
            messagebox.showwarning("提示", "请先加载数据")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("选择分组字段(多选)")
        cols = list(self.data.columns)
        lb = tk.Listbox(dlg, selectmode="multiple", height=10)
        for c in cols:
            lb.insert("end", c)
        lb.pack(padx=12, pady=6)

        def on_ok():
            self.group_batch_fields = [cols[i] for i in lb.curselection()]
            dlg.destroy()

        ttk.Button(dlg, text="确定", command=on_ok, width=10).pack(pady=6)
        dlg.grab_set()
        self.root.wait_window(dlg)

    def _build_buttons(self):
        btnf = ttk.Frame(self.root)
        btnf.pack(pady=12)

        self.calculate_btn = ttk.Button(
            btnf, text="计算捞翔并导出",
            command=self.calculate, width=16
        )
        self.calculate_btn.grid(row=0, column=0, padx=10)

        ttk.Button(
            btnf, text="小捞翔彩蛋",
            command=self.easter_egg, width=10
        ).grid(row=0, column=1)

    def _build_status_bar(self):
        self.progress = ttk.Progressbar(self.root, mode="indeterminate")
        self.progress.pack(fill="x", padx=12, pady=(0, 4))
        self.progress.pack_forget()
        self.status_var = tk.StringVar(value=f"{self.mode_label} 就绪")
        ttk.Label(
            self.root,
            textvariable=self.status_var,
            relief="sunken",
            anchor="w",
        ).pack(side="bottom", fill="x")

    def _round_result_frame(self, df, group_fields, need_area):
        dec = self.area_decimals_var.get()

        def round_dec(value, digits):
            try:
                fmt = "0" if digits == 0 else "0." + "0" * digits
                return float(Decimal(str(value)).quantize(Decimal(fmt), ROUND_HALF_UP))
            except Exception:
                return value

        for column in df.columns:
            if column in group_fields or column in {"数量", "缺失数"}:
                continue
            if column == "面积(亩)" and need_area:
                df[column] = df[column].apply(lambda value: round_dec(value, dec))
            else:
                df[column] = df[column].apply(round2)
        return df

    def _write_export_overview(self, writer, output_path, overview_rows):
        active_filters = [c for c in self.filter_conditions if c.get("enabled", True)]
        filtered_rows = len(self.get_active_data()) if hasattr(self, "get_active_data") else len(self.data)
        summary_rows = [
            {"项目": "导出时间", "内容": time.strftime("%Y-%m-%d %H:%M:%S")},
            {"项目": "模式", "内容": self.mode_label},
            {"项目": "源文件", "内容": self.excel_path},
            {"项目": "数据源", "内容": self.current_sheet_name or "当前数据"},
            {"项目": "原始记录数", "内容": len(self.data)},
            {"项目": "导出记录数", "内容": filtered_rows},
            {"项目": "导出文件", "内容": output_path},
            {"项目": "统计量", "内容": "、".join(self.last_stats)},
            {"项目": "面积统计", "内容": "是" if self.area_cb.get() else "否"},
            {"项目": "筛选条件", "内容": describe_filter_conditions(active_filters)},
            {"项目": "当前模板", "内容": self.active_group_template_name or "未使用"},
            {"项目": "批量值字段", "内容": "、".join(self.batch_fields) if self.batch_fields else "未启用"},
            {"项目": "批量分组字段", "内容": "、".join(self.group_batch_fields) if self.group_batch_fields else "未启用"},
        ]
        pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name="统计说明")
        if overview_rows:
            pd.DataFrame(overview_rows).to_excel(writer, index=False, sheet_name="导出清单")

    def _load_path(self, path):
        self._show_progress("读取文件中...")
        self.excel_path = path
        self.lbl_file.config(text=os.path.basename(path))
        self.add_recent(path)

        if path.lower().endswith(".csv"):
            self.sheet_menu.set("")
            self.sheet_menu["values"] = []
            self.sheet_menu.state(["disabled"])
            self.calculate_btn.state(["disabled"])
            threading.Thread(target=self._load_csv_thread, daemon=True).start()
        else:
            self.calculate_btn.state(["disabled"])
            threading.Thread(target=self._load_excel_sheets_thread, daemon=True).start()

    def _load_csv_thread(self):
        try:
            df = read_csv_safely(self.excel_path)
        except Exception as e:
            message = str(e)
            self.root.after(
                0,
                lambda: (self._hide_progress("就绪"), messagebox.showerror("错误", f"读取失败:\n{message}"))
            )
            return

        self.root.after(0, lambda: self._on_sheet_loaded(df, "CSV"))

    def _load_excel_sheets_thread(self):
        try:
            sheets = pd.ExcelFile(self.excel_path).sheet_names
        except Exception as e:
            message = str(e)
            self.root.after(
                0,
                lambda: (self._hide_progress("就绪"), messagebox.showerror("错误", f"读取失败:\n{message}"))
            )
            return

        def on_loaded():
            self.sheet_menu["values"] = sheets
            self.sheet_menu.state(["!disabled"])
            if sheets:
                self.sheet_var.set(sheets[0])
                self.load_sheet(sheets[0])
            else:
                self._hide_progress("就绪")

        self.root.after(0, on_loaded)

    def select_file(self):
        path = filedialog.askopenfilename(
            title="选择 xlsx 文件",
            filetypes=[("Excel", ("*.xlsx", "*.xls")), ("CSV", "*.csv"), ("所有文件", "*.*")]
        )
        if not path:
            return
        self._load_path(path)

    def handle_drop(self, event):
        path = event.data.strip("{}")
        ext = os.path.splitext(path)[1].lower()
        if ext not in (".xlsx", ".xls", ".csv"):
            messagebox.showerror("错误", "只支持拖拽 .xlsx、.xls 或 .csv 文件")
            return
        self._load_path(path)

    def load_sheet(self, sheet):
        self.sheet_var.set(sheet)
        self.lbl_file.config(text=f"加载中… {sheet}")
        self.calculate_btn.state(["disabled"])
        self._show_progress(f"加载子表: {sheet}")

        threading.Thread(
            target=self._load_sheet_thread,
            args=(sheet,), daemon=True
        ).start()

    def _load_sheet_thread(self, sheet):
        try:
            df = pd.read_excel(self.excel_path, sheet_name=sheet)
        except Exception as e:
            message = str(e)
            self.root.after(
                0,
                lambda: (self._hide_progress("就绪"), messagebox.showerror("错误", f"加载失败:\n{message}"))
            )
            return

        self.root.after(0, lambda: self._on_sheet_loaded(df, sheet))

    def _on_sheet_loaded(self, df, sheet):
        self.data = df
        self.current_sheet_name = sheet
        self.lbl_file.config(text=os.path.basename(self.excel_path))

        cols = list(df.columns)
        # 刷新分类级下拉
        for idx, lvl in enumerate(self.levels):
            var, combo = lvl["var"], lvl["combo"]
            combo["values"] = cols
            if idx < len(cols):
                var.set(cols[idx])

        # 刷新值字段 & 面积字段
        for var, combo in (
                (self.val_var, self.val_menu),
                (self.ratio_var, self.ratio_menu)
        ):
            combo["values"] = cols

        if cols:
            self.val_var.set(cols[-1])
            self.ratio_var.set(cols[-1])
        self._restore_ui_state(cols)
        self._normalize_filter_conditions(cols)
        self.calculate_btn.state(["!disabled"])
        self._hide_progress()
        self._update_loaded_status()

    def add_level(self):
        if len(self.levels) >= MAX_LEVELS:
            return

        var = tk.StringVar()
        lbl = ttk.Label(
            self.frame_levels,
            text=f"分类级 {len(self.levels) + 1}:"
        )
        combo = ttk.Combobox(
            self.frame_levels, textvariable=var,
            state="readonly", width=20
        )
        btn = ttk.Button(self.frame_levels, text="×", width=2)
        btn.config(command=partial(self._remove_level, btn))

        self.levels.append({
            "var": var,
            "label": lbl,
            "combo": combo,
            "btn": btn
        })

        if not self.data.empty:
            combo["values"] = list(self.data.columns)
            var.set(self.data.columns[0])

        self._layout_levels()

    def _remove_level(self, btn):
        for idx, lvl in enumerate(self.levels):
            if lvl["btn"] is btn:
                lvl["label"].destroy()
                lvl["combo"].destroy()
                lvl["btn"].destroy()
                self.levels.pop(idx)
                break

        self._layout_levels()

    def _layout_levels(self):
        for i, lvl in enumerate(self.levels):
            lvl["label"].grid(row=i, column=0, padx=5, pady=2, sticky="w")
            lvl["combo"].grid(
                row=i, column=1, padx=5, pady=2, sticky="w"
            )
            lvl["btn"].grid(row=i, column=2, padx=5)

        self.add_btn.grid(
            row=len(self.levels), column=0,
            columnspan=3, pady=6, sticky="w"
        )

    def choose_stats(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("选择统计量")
        dlg.geometry("560x360")

        vars_ = {}
        for i, stat in enumerate(self.all_stats):
            var = tk.BooleanVar(value=(stat in self.last_stats))
            chk = ttk.Checkbutton(dlg, text=stat, variable=var)
            chk.grid(row=i // 4, column=i % 4, padx=8, pady=5, sticky="w")
            vars_[stat] = var

        ttk.Button(
            dlg, text="确定", width=12,
            command=dlg.destroy
        ).grid(
            row=(len(self.all_stats) + 3) // 4,
            column=0, columnspan=4, pady=12
        )

        dlg.transient(self.root)
        dlg.grab_set()
        self.root.wait_window(dlg)

        self.last_stats = [
            s for s, v in vars_.items() if v.get()
        ]
        return self.last_stats
    def open_stats_order(self):
        """自定义统计量顺序（支持拖拽 + 上下移）"""
        dlg = tk.Toplevel(self.root)
        dlg.title("自定义统计量顺序")
        dlg.geometry("300x400")

        ttk.Label(dlg, text="拖拽或用按钮调整顺序：").pack(pady=6)
        lb = tk.Listbox(dlg)
        lb.pack(fill="both", expand=True, padx=12, pady=6)
        for st in self.last_stats:
            lb.insert("end", st)

        # 拖拽排序支持
        def on_start(e):
            lb._drag_index = lb.nearest(e.y)
        def on_motion(e):
            i = lb.nearest(e.y)
            if i != lb._drag_index:
                v = lb.get(lb._drag_index)
                lb.delete(lb._drag_index)
                lb.insert(i, v)
                lb.select_clear(0, "end")
                lb.select_set(i)
                lb._drag_index = i
        lb.bind("<Button-1>", on_start)
        lb.bind("<B1-Motion>", on_motion)

        # 上/下 移动
        btnf = ttk.Frame(dlg)
        btnf.pack(pady=4)
        def move_up():
            sel = lb.curselection()
            if not sel or sel[0]==0: return
            i=sel[0]; v=lb.get(i)
            lb.delete(i); lb.insert(i-1, v); lb.select_set(i-1)
        def move_down():
            sel = lb.curselection()
            if not sel or sel[0]==lb.size()-1: return
            i=sel[0]; v=lb.get(i)
            lb.delete(i); lb.insert(i+1, v); lb.select_set(i+1)
        ttk.Button(btnf, text="上移",   command=move_up).grid(row=0,column=0,padx=6)
        ttk.Button(btnf, text="下移",   command=move_down).grid(row=0,column=1,padx=6)

        def on_ok():
            # 更新顺序
            self.last_stats = [lb.get(i) for i in range(lb.size())]
            dlg.destroy()
        ttk.Button(dlg, text="确定", command=on_ok, width=10).pack(pady=(0,12))

        dlg.grab_set()
        self.root.wait_window(dlg)

    def calculate(self):
        start_time = time.time()

        if self.data.empty:
            messagebox.showerror("错误", "请先选择并加载数据")
            return

        active_df = self._get_filtered_data(show_error=True)
        if active_df is None:
            return
        if active_df.empty:
            messagebox.showwarning("提示", "筛选后没有可用数据，请调整条件后再试。")
            return

        # 直接使用按钮预先设置好的统计量顺序
        stats = self.last_stats.copy()
        if not stats:
            messagebox.showwarning("提示", "未选择任何统计量，已取消")
            return
        # 确定分组字段
        if self.group_batch_var.get() and self.group_batch_fields:
            group_fields_list = [[field] for field in self.group_batch_fields]
        else:
            group_fields = [
                lvl["var"].get() for lvl in self.levels
                if lvl["var"].get()
            ]
            group_fields_list = [group_fields] if group_fields else [[]]

        # 确定值字段
        if self.batch_var.get() and self.batch_fields:
            val_fields = self.batch_fields
        else:
            val_fields = [self.val_var.get()]

        ratio = self.ratio_var.get()
        need_area = self.area_cb.get()

        # 校验字段
        all_fields = set()
        for group_fields in group_fields_list:
            all_fields.update(group_fields)
        all_fields.update(val_fields)
        if need_area:
            all_fields.add(ratio)

        for f in all_fields:
            if f not in active_df.columns:
                messagebox.showerror("错误", f"列 '{f}' 不存在")
                return

        self.calculate_btn.state(["disabled"])
        self._show_progress("正在计算...")

        # 准备输出文件
        base = os.path.splitext(os.path.basename(self.excel_path))[0]
        out = build_output_path(self, f"{base}_计算小捞翔后")

        overview_rows = []
        skipped = []
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            for val in val_fields:
                for group_fields in group_fields_list:
                    df0 = active_df.copy()
                    df0[val] = pd.to_numeric(df0[val], errors="coerce")
                    valid_count = int(df0[val].count())
                    if valid_count == 0:
                        skipped.append((val, group_fields))
                        continue

                    if need_area:
                        df0[ratio] = pd.to_numeric(df0[ratio], errors="coerce")
                        if group_fields:
                            area_total = df0.groupby(group_fields, dropna=False)[ratio].sum().reset_index(name="面积(亩)")
                        else:
                            area_total = None

                    df = build_grouped_stats_frame(df0, val, group_fields)

                    if need_area:
                        if group_fields:
                            df = df.merge(area_total, on=group_fields, how="left")
                            total_area = df["面积(亩)"].sum()
                        else:
                            total_area = float(df0[ratio].sum())
                            df["面积(亩)"] = total_area
                        df["面积占比"] = (df["面积(亩)"] / total_area * 100) if total_area else np.nan

                    df = self._round_result_frame(df, group_fields, need_area)
                    df = apply_count_masks(df)

                    # 排序
                    if group_fields:
                        for col in reversed(group_fields):
                            if col in self.custom_orders and self.custom_orders[col]:
                                present = df[col].dropna().unique().tolist()
                                seq = [v for v in self.custom_orders[col] if v in present]

                                if seq:
                                    if pd.api.types.is_numeric_dtype(df[col]):
                                        try:
                                            seq = [float(v) if isinstance(v, str) and v.replace('.', '',
                                                                                                1).isdigit() else v for
                                                   v in seq]
                                        except:
                                            pass

                                    df[col] = pd.Categorical(df[col], categories=seq, ordered=True)
                                    df = df.sort_values(by=col, kind="stable")
                                else:
                                    df = df.sort_values(
                                        by=col,
                                        key=lambda s: s.map(sort_key),
                                        kind="stable"
                                    )
                            else:
                                df = df.sort_values(
                                    by=col,
                                    key=lambda s: s.map(sort_key),
                                    kind="stable"
                                )

                    if need_area:
                        current_stats = stats + ["面积(亩)", "面积占比"]
                    else:
                        current_stats = stats.copy()

                    mlt = df.melt(
                        id_vars=group_fields,
                        value_vars=current_stats,
                        var_name="指标",
                        value_name="数值"
                    )

                    # 强制指标顺序
                    order = {
                        s: i for i, s in enumerate(
                            stats + (["面积(亩)", "面积占比"] if need_area else [])
                        )
                    }
                    mlt["__ord"] = mlt["指标"].map(order)
                    mlt.sort_values(group_fields + ["__ord"], inplace=True)
                    mlt.drop(columns="__ord", inplace=True)

                    # 生成sheet名称
                    sheet_name = f"{val}"
                    if group_fields:
                        sheet_name += f"_{'_'.join(group_fields)}"
                    sheet_name = sheet_name[:31]  # Excel sheet名称长度限制

                    # 写入sheet
                    mlt.to_excel(writer, index=False, sheet_name=sheet_name)
                    ws = writer.sheets[sheet_name]
                    overview_rows.append({
                        "工作表": sheet_name,
                        "值字段": val,
                        "分组字段": " / ".join(group_fields) if group_fields else "无分组",
                        "导出行数": len(mlt),
                        "有效数量": valid_count,
                        "面积统计": "是" if need_area else "否",
                    })

                    # 合并分类列
                    if group_fields:
                        start_row = 2
                        total_rows = len(mlt)

                        for ci in range(len(group_fields)):
                            prev = ws.cell(row=start_row, column=ci + 1).value
                            ms = start_row

                            for off in range(1, total_rows):
                                r = start_row + off
                                curr = ws.cell(row=r, column=ci + 1).value

                                if curr != prev:
                                    if r - ms >= 2:
                                        ws.merge_cells(
                                            start_row=ms,
                                            start_column=ci + 1,
                                            end_row=r - 1,
                                            end_column=ci + 1
                                        )
                                    prev, ms = curr, r

                            last = start_row + total_rows - 1
                            if last - ms >= 1:
                                ws.merge_cells(
                                    start_row=ms,
                                    start_column=ci + 1,
                                    end_row=last,
                                    end_column=ci + 1
                                )

                    # 设置居中 + 边框
                    thin = Side(border_style="thin", color="000000")
                    border = Border(
                        top=thin, bottom=thin,
                        left=thin, right=thin
                    )

                    for row in ws.iter_rows(
                            min_row=1,
                            max_row=ws.max_row,
                            min_col=1,
                            max_col=ws.max_column
                    ):
                        for cell in row:
                            cell.alignment = Alignment(
                                horizontal="center",
                                vertical="center"
                            )
                            cell.border = border

            self._write_export_overview(writer, out, overview_rows)

        if not overview_rows:
            if os.path.exists(out):
                os.remove(out)
            messagebox.showerror("错误", "没有找到可用于统计的数值数据。")
            self.calculate_btn.state(["!disabled"])
            return self._hide_progress("就绪")

        if skipped:
            skipped_text = "\n".join(
                f"- 值字段 {field} / 分组 {' / '.join(groups) if groups else '无分组'}"
                for field, groups in skipped[:8]
            )
            messagebox.showwarning("部分组合已跳过", f"以下组合没有有效数值，已自动跳过：\n{skipped_text}")

        elapsed = round2(time.time() - start_time)
        self.calculate_btn.state(["!disabled"])
        self._hide_progress(f"完成: {elapsed}s")
        self.show_result_dialog(out, elapsed)

    def _show_progress(self, text):
        begin_busy(self, text)

    def _hide_progress(self, text=None):
        end_busy(self, text)

    def easter_egg(self):
        """
        彩蛋：生成 1～999 个随机 Excel 文件
        """
        for i in range(1, 1000):
            fn = f"{i}小捞翔.xlsx"
            df = pd.DataFrame(
                np.random.randint(0, 10 ** 9, size=(500, 100)))
            df.to_excel(fn, index=False)

    def open_custom_sort(self):
        """与横版相同的 自定义排序 对话框"""
        # 收集所有分组列
        cols = [lvl["var"].get() for lvl in self.levels if lvl["var"].get()]

        if not cols:
            messagebox.showwarning("提示", "请先设置分类级别")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("自定义排序")
        dlg.geometry("300x440")

        ttk.Label(dlg, text="请选择列:").pack(pady=(10, 0))
        col_var = tk.StringVar(value=cols[0])
        cmb = ttk.Combobox(dlg, values=cols, textvariable=col_var,
                           state="readonly")
        cmb.pack(fill="x", padx=12, pady=6)

        lb = tk.Listbox(dlg)
        lb.pack(fill="both", expand=True, padx=12, pady=6)
        last_col = None

        def _on_start(e):
            lb._drag_index = lb.nearest(e.y)

        def _on_motion(e):
            i = lb.nearest(e.y)
            if i != lb._drag_index:
                v = lb.get(lb._drag_index)
                lb.delete(lb._drag_index)
                lb.insert(i, v)
                lb.select_clear(0, "end")
                lb.select_set(i)
                lb._drag_index = i

        lb.bind("<Button-1>", _on_start)
        lb.bind("<B1-Motion>", _on_motion)

        # 上/下 移动按钮
        btnf = ttk.Frame(dlg);
        btnf.pack(pady=4)

        def up():
            sel = lb.curselection()
            if not sel or sel[0] == 0: return
            i = sel[0];
            v = lb.get(i)
            lb.delete(i);
            lb.insert(i - 1, v);
            lb.select_set(i - 1)

        def down():
            sel = lb.curselection()
            if not sel or sel[0] == lb.size() - 1: return
            i = sel[0];
            v = lb.get(i)
            lb.delete(i);
            lb.insert(i + 1, v);
            lb.select_set(i + 1)

        ttk.Button(btnf, text="上移", command=up).grid(row=0, column=0, padx=6)
        ttk.Button(btnf, text="下移", command=down).grid(row=0, column=1, padx=6)

        def load_vals(e=None):
            nonlocal last_col
            # 保存当前列的排序设置
            if last_col:
                self.custom_orders[last_col] = list(lb.get(0, 'end'))

            col = col_var.get()
            last_col = col

            # 如果该列还没有原始顺序记录，则创建
            if col not in self.original_orders:
                vals = list(self.data[col].dropna().unique())
                self.original_orders[col] = vals.copy()

            lb.delete(0, 'end')

            # 优先使用自定义排序，如果没有则使用原始顺序
            if col in self.custom_orders and self.custom_orders[col]:
                for v in self.custom_orders[col]:
                    lb.insert('end', v)
            else:
                for v in self.original_orders[col]:
                    lb.insert('end', v)

        load_vals()
        cmb.bind("<<ComboboxSelected>>", load_vals)

        # 导入 / 恢复 / 导出 按钮区
        tools = ttk.Frame(dlg);
        tools.pack(pady=(0, 8))

        def import_txt():
            path = filedialog.askopenfilename(
                title="导入排序 TXT",
                filetypes=[("Text 文件", "*.txt"), ("All", "*.*")]
            )
            if not path: return
            lines = [l.strip() for l in open(path, encoding='utf-8') if l.strip()]
            curr = list(lb.get(0, 'end'))
            if len(lines) != len(curr) or set(lines) != set(curr):
                if not messagebox.askyesno("警告",
                                           f"行数 {len(lines)} 与当前值 {len(curr)} 不匹配，继续？"):
                    return
            lb.delete(0, 'end')
            for v in lines: lb.insert('end', v)

        def restore():
            col = col_var.get()
            lb.delete(0, 'end')
            for v in self.original_orders.get(col, []):
                lb.insert('end', v)

        ttk.Button(tools, text="从 TXT 导入", command=import_txt).grid(row=0, column=0, padx=6)
        ttk.Button(tools, text="恢复原始顺序", command=restore).grid(row=0, column=1, padx=6)

        def export_txt():
            path = filedialog.asksaveasfilename(
                title="导出排序 TXT",
                defaultextension=".txt",
                filetypes=[("Text 文件", "*.txt"), ("All", "*.*")]
            )
            if not path: return
            with open(path, 'w', encoding='utf-8') as f:
                for v in lb.get(0, 'end'):
                    f.write(f"{v}\n")
            messagebox.showinfo("导出成功", f"已导出到:\n{path}")

        ttk.Button(tools, text="导出排序", command=export_txt).grid(row=0, column=2, padx=6)

        def on_ok():
            # 只保存当前列的排序设置
            if last_col:
                self.custom_orders[last_col] = list(lb.get(0, 'end'))
            dlg.destroy()

        ttk.Button(dlg, text="确定", command=on_ok, width=10).pack(pady=(0, 12))

        dlg.grab_set()
        self.root.wait_window(dlg)
    def show_result_dialog(self, filepath, elapsed):
        mark_output(self, filepath)
        self.save_config(show_msg=False)
        dlg = tk.Toplevel(self.root)
        dlg.title("完成")
        dlg.geometry("620x190")
        dlg.resizable(False, False)

        msg = f"已导出文件：\n{filepath}\n\n耗时 {elapsed} 秒"
        ttk.Label(dlg, text=msg, justify="left").pack(padx=12, pady=(12,4))

        frm = ttk.Frame(dlg)
        frm.pack(pady=(0,12))
        ttk.Button(frm, text="打开文件",
                   command=lambda: self.open_file(filepath))\
            .grid(row=0, column=0, padx=6)
        ttk.Button(frm, text="预览",
                   command=lambda: self.open_preview(filepath))\
            .grid(row=0, column=1, padx=6)
        ttk.Button(frm, text="打开文件夹",
                   command=lambda: open_output_folder(self))\
            .grid(row=0, column=2, padx=6)
        ttk.Button(frm, text="复制路径",
                   command=lambda: copy_text(self, filepath, "导出路径已复制"))\
            .grid(row=0, column=3, padx=6)
        ttk.Button(frm, text="关闭",
                   command=dlg.destroy)\
            .grid(row=0, column=4, padx=6)

        dlg.grab_set()

    def open_file(self, filepath):
        try:
            open_with_default_app(filepath)
        except Exception as e:
            messagebox.showerror("打开失败", str(e))

    def open_preview(self, filepath):
        show_workbook_preview(self.root, filepath, title="预览结果", default_limit=300)
