import itertools
import json
import os
import threading
import time
import traceback
from decimal import Decimal, ROUND_HALF_UP

import numpy as np
import pandas as pd
import tkinter as tk
from openpyxl.styles import Alignment, Border, Side
from tkinter import filedialog, messagebox, simpledialog, ttk
from tkinterdnd2 import DND_FILES

from common_utils import (
    EXTENDED_STATS,
    FILTER_OPERATORS,
    apply_filter_conditions,
    apply_count_masks,
    build_grouped_stats_frame,
    calculate_series_stats,
    describe_filter_conditions,
    open_with_default_app,
    read_csv_safely,
    round2,
    sort_key,
)
from ui_shell import (
    bind_shortcuts,
    build_output_path,
    build_app_menu,
    copy_text,
    initialize_shell,
    mark_output,
    maybe_restore_recent_file,
    open_output_folder,
    populate_recent_menus,
    set_status,
    show_workbook_preview,
)

# 当前程序版本号——每次发布时请手动更新
APP_VERSION = "1.2.0"

UPDATE_CONTENT = """
小捞翔·至尊版 v1.2.0 更新日志：
- 新增：扩展统计指标，支持缺失数、缺失率、四分位数、极差、方差、偏度、峰度等
- 新增：横版支持“插入小计行”开关，统计项里的“合计”现在会作为求和列正常导出
- 新增：自动生成“统计说明”和“导出清单”工作表，方便回看参数和输出内容
- 新增：值字段会自动尝试转为数值，无法计算的组合会提示并自动跳过
- 优化：横版和竖版统计口径统一，更多统计方法在两个界面保持一致
- 优化：统计量选择面板扩大，容纳更多指标
"""


MAX_LEVELS = 3
# 默认配置文件名（可自定义目录）
CONFIG_FILE = "small_shit.json"

# 1) 定义一个可折叠面板类

class HorizontalApp:
    def __init__(self, root):
        self.root = root
        initialize_shell(self, mode_name="horizontal", title="小捞翔")
        root.geometry("780x720")
        root.minsize(700, 620)
        root.resizable(True, True)

        # 在窗口关闭前保存配置
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        root.drop_target_register(DND_FILES)
        root.dnd_bind('<<DragEnter>>',  lambda e: root.configure(bg='#e3f2fd'))
        root.dnd_bind('<<DragLeave>>',  lambda e: root.configure(bg='SystemButtonFace'))
        root.dnd_bind('<<Drop>>',       self.handle_drop)

        def sep(): ttk.Separator(root, orient='horizontal').pack(fill='x', pady=6)

        self.excel_path = ""
        # … 其他属性初始化 …
        self.recent_files = [] 
        self.custom_orders = {}    # 用户确认后的自定义顺序
        self.original_orders = {}
        self.data = pd.DataFrame()
        self.all_stats = EXTENDED_STATS.copy()
        self.last_stats = self.all_stats.copy()
        self.batch_fields = []
        self.group_batch_fields = []
        self.filter_conditions = []
        self.group_templates = {}
        self.active_group_template_name = ""

        # 文件 & Sheet 选择
        frm = ttk.Frame(root); frm.pack(fill='x', padx=12, pady=6)
        ttk.Button(frm, text="选择 Excel", command=self.select_file)\
            .grid(row=0, column=0, padx=(0,8))
        self.lbl_file = ttk.Label(frm, text="— 未选择文件 —")
        self.lbl_file.grid(row=0, column=1, sticky='w')
        ttk.Label(frm, text="子表:").grid(row=1, column=0, pady=6)
        self.sheet_var = tk.StringVar()
        self.sheet_menu = ttk.Combobox(frm, textvariable=self.sheet_var,
                                       state="readonly", width=38)
        self.sheet_menu.grid(row=1, column=1, sticky='w')
        self.sheet_menu.bind("<<ComboboxSelected>>",
                             lambda e: self.load_sheet(self.sheet_var.get()))
        sep()
        # “选择 Excel”旁边的“最近打开”下拉
        self.recent_btn = ttk.Menubutton(frm, text="最近打开")
        self.recent_menu = tk.Menu(self.recent_btn, tearoff=0)
        self.recent_btn["menu"] = self.recent_menu
        self.recent_btn.grid(row=0, column=2, padx=(8, 0))

        # 分类级别(1~3级)
        self.frame_levels = ttk.LabelFrame(root, text="分类级别(1~3级)")
        self.frame_levels.pack(fill='x', padx=12, pady=6)

        self.levels = []
        self.add_btn = ttk.Button(self.frame_levels, text="+ 添加级别",
                                  command=self.add_level)
        self.add_level(); self.add_level()
        sep()


        # 值字段 & 面积占比
        frm2 = ttk.Frame(root)
        frm2.pack(fill='x', padx=12, pady=6)

        ttk.Label(frm2, text="值字段:").grid(row=0, column=0)
        self.val_var = tk.StringVar()
        self.val_menu = ttk.Combobox(frm2, textvariable=self.val_var,
                                     state="readonly", width=36)
        self.val_menu.grid(row=0, column=1, sticky='w')
        self.val_enable_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(frm2, text="启用值字段",
                        variable=self.val_enable_var,
                        command=self.toggle_value_field) \
            .grid(row=0, column=2, padx=6)
        ttk.Label(frm2, text="面积字段:").grid(row=1, column=0, pady=6)
        self.ratio_var = tk.StringVar()
        self.ratio_menu = ttk.Combobox(frm2, textvariable=self.ratio_var,
                                       state="readonly", width=36)
        self.ratio_menu.grid(row=1, column=1, sticky='w')
        self.area_cb = tk.BooleanVar()
        # … 已有的面积字段 & 复选框 …
        ttk.Checkbutton(frm2, text="计算面积占比", variable=self.area_cb,
                        command=self.toggle_area_fields) \
            .grid(row=1, column=2, padx=6)

        # 新增：面积小数位选择
        ttk.Label(frm2, text="面积小数位:").grid(row=2, column=0, pady=(4, 0))
        self.area_decimals_var = tk.IntVar(value=2)
        self.area_decimals_spin = ttk.Spinbox(
            frm2, from_=0, to=10, textvariable=self.area_decimals_var,
            width=5, state="readonly"
        )
        self.area_decimals_spin.grid(row=2, column=1, sticky='w')

        sep()

        # 统计量 & 批量 & 自定义排序

        # 一定要先定义这两个属性
        self.batch_var       = tk.BooleanVar()
        self.group_batch_var = tk.BooleanVar()
        self.subtotal_var = tk.BooleanVar(value=False)

        # 统计量 & 批量 & 自定义排序 分两行布局
        stats_frm = ttk.Frame(root)
        stats_frm.pack(fill='x', padx=12, pady=6)

        # 第一行：统计量
        row1 = ttk.Frame(stats_frm)
        row1.pack(fill='x', pady=2)
        ttk.Button(row1, text="选择统计量", command=self.choose_stats, width=16)\
            .pack(side='left', padx=6)
        ttk.Button(row1, text="自定义统计量顺序", command=self.open_stats_order, width=16)\
            .pack(side='left', padx=6)
        ttk.Button(row1, text="自定义排序", command=self.open_custom_sort, width=10)\
            .pack(side='left', padx=6)
        ttk.Button(row1, text="条件筛选", command=self.open_filter_builder, width=10)\
            .pack(side='left', padx=6)
        ttk.Button(row1, text="分组模板", command=self.open_group_template_manager, width=10)\
            .pack(side='left', padx=6)
        # 第二行：批量 & 分组 & 自定义排序
        row2 = ttk.Frame(stats_frm)
        row2.pack(fill='x', pady=2)
        ttk.Checkbutton(
            row2, text="批量值字段", variable=self.batch_var,
            command=self.toggle_batch
        ).pack(side='left', padx=6)
        self.batch_btn = ttk.Button(
            row2, text="选择值字段",
            command=self.choose_batch_fields,
            state="disabled", width=10
        )
        self.batch_btn.pack(side='left', padx=6)

        ttk.Checkbutton(
            row2, text="批量分组(级别1)",
            variable=self.group_batch_var,
            command=self.toggle_group_batch
        ).pack(side='left', padx=6)
        self.group_batch_btn = ttk.Button(
            row2, text="选择分组字段",
            command=self.choose_group_batch_fields,
            state="disabled", width=12
        )
        self.group_batch_btn.pack(side='left', padx=6)
        ttk.Checkbutton(
            row2, text="插入小计行", variable=self.subtotal_var
        ).pack(side='left', padx=6)
        sep()

        # 操作按钮
        btnf = ttk.Frame(root); btnf.pack(pady=8)
        self.calculate_btn = ttk.Button(btnf, text="计算并导出",
                                        command=self.calculate, width=16)
        self.calculate_btn.grid(row=0, column=0, padx=6)
        ttk.Button(btnf, text="彩蛋", command=self.easter_egg,
                   width=8).grid(row=0, column=1)
        sep()

        # 进度条 & 状态栏
        self.progress = ttk.Progressbar(root, mode="indeterminate")
        self.progress.pack(fill='x', padx=12); self.progress.pack_forget()
        self.status_var = tk.StringVar(value=f"{self.mode_label} 就绪")
        ttk.Label(root, textvariable=self.status_var, relief='sunken',
                  anchor='w').pack(side='bottom', fill='x')

        self.toggle_area_fields()
        # 载入上次保存的配置
        self.load_config()
        # 确保存在默认结构
        self.advanced_order = getattr(self, "advanced_order",
                                      {"groups": [], "stats": [], "exports": []})

        build_app_menu(self)
        bind_shortcuts(self)
        self.create_recent_menu()
        self.check_for_update()
        maybe_restore_recent_file(self)
    def show_update_history(self):
        history = getattr(self, "update_history", [])
        if not history:
            messagebox.showinfo("更新历史", "暂无更新记录。")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("更新历史")
        dlg.geometry("520x360")
        dlg.resizable(False, False)

        frm = ttk.Frame(dlg, padding=8)
        frm.pack(fill="both", expand=True)

        txt = tk.Text(frm, wrap="word")
        # 一条一条字体换行
        txt.insert("1.0", "\n\n".join(history))
        txt.configure(state="disabled")
        txt.pack(fill="both", expand=True)

        btn = ttk.Button(frm, text="关闭", command=dlg.destroy)
        btn.pack(pady=6)

        dlg.transient(self.root)
        dlg.grab_set()
        self.root.wait_window(dlg)

    def check_for_update(self):
        if getattr(self, "shown_version", "") != APP_VERSION:
            dlg = tk.Toplevel(self.root)
            dlg.title(f"更新日志  v{APP_VERSION}")
            dlg.geometry("520x360")
            dlg.resizable(False, False)

            # — 文本区 —
            frm_txt = ttk.Frame(dlg, padding=8)
            frm_txt.pack(fill="both", expand=True)
            txt = tk.Text(frm_txt, wrap="word", height=10)
            txt.insert("1.0", UPDATE_CONTENT.strip())
            txt.configure(state="disabled")  # 只读
            txt.pack(fill="both", expand=True)

            # — 按钮区 —
            frm_btn = ttk.Frame(dlg, padding=(8, 8))
            frm_btn.pack(fill="x")

            def _on_close():
                self.shown_version = APP_VERSION
                # 把本次更新内容加入历史（避免重复）
                if UPDATE_CONTENT.strip() not in self.update_history:
                    self.update_history.append(UPDATE_CONTENT.strip())
                self.save_config(show_msg=False)

                dlg.destroy()

            btn = ttk.Button(frm_btn, text="知道了", command=_on_close)
            # 把按钮靠右放，有内边距
            btn.pack(side="right", padx=4, pady=4)

            dlg.transient(self.root)
            dlg.grab_set()
            self.root.wait_window(dlg)

    # 批量值字段
    # 自定义排序对话框：选择分组列，拖拽/上下移动唯一值
    def save_config(self, path=None, show_msg=True):

        cfg = {
            "custom_orders": self.custom_orders,
            "last_stats": self.last_stats,
            "recent_files": self.recent_files,
            "shown_version": getattr(self, "shown_version", ""),
            "update_history": getattr(self, "update_history", []),
            "area_decimals": self.area_decimals_var.get(),
            "ui_state": self._collect_ui_state(),
            "filter_conditions": self.filter_conditions,
            "group_templates": self.group_templates,
            "active_group_template_name": self.active_group_template_name,
            "last_output_path": self.last_output_path,
        }
        cfg_path = path or CONFIG_FILE
        try:
            with open(cfg_path, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
            if show_msg:
                messagebox.showinfo("配置已保存", f"配置已保存到：\n{cfg_path}")
        except Exception as e:
            if show_msg:
                messagebox.showerror("保存失败", str(e))

    def load_config(self, path=None):
        # 确保 update_history 属性总是存在
        self.update_history = []
        """
        从 JSON 配置文件加载：custom_orders, last_stats, recent_files,
        shown_version, update_history 以及 area_decimals（面积小数位）。
        """
        cfg_path = path or CONFIG_FILE
        # 如果配置文件不存在，直接返回（保持各属性的默认值）
        if not os.path.exists(cfg_path):
            return

        try:
            with open(cfg_path, "r", encoding="utf-8") as f:
                cfg = json.load(f)

            # 用户在 自定义排序 中保存的列顺序
            self.custom_orders = cfg.get("custom_orders", {})

            # 上次选择的统计量列表
            self.last_stats = cfg.get("last_stats", self.all_stats.copy())

            # 最近打开文件列表
            self.recent_files = cfg.get("recent_files", [])

            # 已展示过更新提示的版本号
            self.shown_version = cfg.get("shown_version", "")

            # 已累计的更新日志历史（多条）
            self.update_history = cfg.get("update_history", [])

            # 面积小数位：默认为已有控件的当前值
            dec = cfg.get("area_decimals", self.area_decimals_var.get())
            self.area_decimals_var.set(dec)
            self.saved_ui_state = cfg.get("ui_state", {})
            self.filter_conditions = cfg.get("filter_conditions", [])
            self.group_templates = cfg.get("group_templates", {})
            self.active_group_template_name = cfg.get("active_group_template_name", "")
            self.last_output_path = cfg.get("last_output_path")

        except Exception as e:
            messagebox.showerror("加载失败", str(e))

    def on_closing(self):
        # 退出前保存配置
        self.save_config(show_msg=False)
        self.root.destroy()

    def _collect_ui_state(self):
        return {
            "levels": [lvl["var"].get() for lvl in self.levels if lvl["var"].get()],
            "value_field": self.val_var.get(),
            "ratio_field": self.ratio_var.get(),
            "value_enabled": bool(self.val_enable_var.get()),
            "area_enabled": bool(self.area_cb.get()),
            "batch_enabled": bool(self.batch_var.get()),
            "batch_fields": list(self.batch_fields),
            "group_batch_enabled": bool(self.group_batch_var.get()),
            "group_batch_fields": list(self.group_batch_fields),
            "subtotal_enabled": bool(self.subtotal_var.get()),
        }

    def _set_level_count(self, target_count):
        target_count = max(1, min(MAX_LEVELS, target_count))
        while len(self.levels) < target_count:
            self.add_level()
        while len(self.levels) > target_count:
            self._remove_level(self.levels[-1]["btn"])

    def _restore_ui_state(self, cols):
        state = getattr(self, "saved_ui_state", {})
        if not isinstance(state, dict):
            return

        saved_levels = [field for field in state.get("levels", []) if field in cols]
        if saved_levels:
            self._set_level_count(len(saved_levels))
            for idx, lvl in enumerate(self.levels):
                if idx < len(saved_levels):
                    lvl["var"].set(saved_levels[idx])
                elif cols:
                    lvl["var"].set(cols[0])

        saved_value = state.get("value_field")
        saved_ratio = state.get("ratio_field")
        if saved_value in cols:
            self.val_var.set(saved_value)
        if saved_ratio in cols:
            self.ratio_var.set(saved_ratio)

        self.val_enable_var.set(bool(state.get("value_enabled", self.val_enable_var.get())))
        self.toggle_value_field()
        if self.val_enable_var.get() and saved_value in cols:
            self.val_var.set(saved_value)

        self.area_cb.set(bool(state.get("area_enabled", self.area_cb.get())))
        self.toggle_area_fields()
        if self.area_cb.get() and saved_ratio in cols:
            self.ratio_var.set(saved_ratio)

        self.batch_var.set(bool(state.get("batch_enabled", False)))
        self.toggle_batch()
        self.batch_fields = [field for field in state.get("batch_fields", []) if field in cols]

        self.group_batch_var.set(bool(state.get("group_batch_enabled", False)))
        self.toggle_group_batch()
        self.group_batch_fields = [field for field in state.get("group_batch_fields", []) if field in cols]
        self.subtotal_var.set(bool(state.get("subtotal_enabled", self.subtotal_var.get())))

    def _normalize_filter_conditions(self, cols):
        normalized = []
        for condition in getattr(self, "filter_conditions", []):
            if not isinstance(condition, dict):
                continue
            field = str(condition.get("field", "")).strip()
            op = str(condition.get("op", "")).strip()
            if field and op and field in cols:
                normalized.append(
                    {
                        "field": field,
                        "op": op,
                        "value": str(condition.get("value", "")),
                        "enabled": bool(condition.get("enabled", True)),
                    }
                )
        self.filter_conditions = normalized

    def _filter_summary(self):
        active_conditions = [c for c in self.filter_conditions if c.get("enabled", True)]
        return describe_filter_conditions(active_conditions)

    def _get_filtered_data(self, show_error=True):
        if self.data.empty:
            return self.data.copy()

        try:
            return apply_filter_conditions(self.data, self.filter_conditions)
        except Exception as exc:
            if show_error:
                messagebox.showerror("筛选条件有误", str(exc))
                set_status(self, "筛选条件有误，请检查后再试")
            return None

    def get_active_data(self):
        filtered = self._get_filtered_data(show_error=False)
        if isinstance(filtered, pd.DataFrame):
            return filtered
        return self.data

    def _update_loaded_status(self):
        if self.data.empty:
            set_status(self, f"{self.mode_label} 就绪")
            return

        total_rows = len(self.data)
        total_cols = len(self.data.columns)
        if not self.filter_conditions:
            set_status(self, f"已加载 {total_rows} 行 / {total_cols} 列")
            return

        filtered = self._get_filtered_data(show_error=False)
        if filtered is None:
            set_status(self, f"已加载 {total_rows} 行 / {total_cols} 列，筛选条件待修正")
            return

        set_status(self, f"已加载 {total_rows} 行 / {total_cols} 列，筛选后 {len(filtered)} 行")

    def _collect_group_template(self):
        return {
            "ui_state": self._collect_ui_state(),
            "stats": list(self.last_stats),
        }

    def _describe_group_template(self, name, payload):
        state = payload.get("ui_state", {}) if isinstance(payload, dict) else {}
        stats = payload.get("stats", []) if isinstance(payload, dict) else []
        lines = [
            f"模板名称：{name}",
            f"分类级别：{' / '.join(state.get('levels', [])) or '未设置'}",
            f"值字段：{state.get('value_field') or '未设置'}",
            f"面积字段：{state.get('ratio_field') or '未设置'}",
            f"启用值字段：{'是' if state.get('value_enabled', True) else '否'}",
            f"面积统计：{'是' if state.get('area_enabled', False) else '否'}",
            f"批量值字段：{' / '.join(state.get('batch_fields', [])) or '未启用'}",
            f"批量分组字段：{' / '.join(state.get('group_batch_fields', [])) or '未启用'}",
            f"插入小计行：{'是' if state.get('subtotal_enabled', False) else '否'}",
            f"统计量：{' / '.join(stats) if stats else '未设置'}",
        ]
        return "\n".join(lines)

    def _apply_group_template(self, payload, template_name=""):
        if self.data.empty:
            messagebox.showwarning("提示", "请先加载数据后再应用模板")
            return False

        if not isinstance(payload, dict):
            messagebox.showerror("错误", "模板内容无效")
            return False

        state = payload.get("ui_state", {})
        if not isinstance(state, dict):
            messagebox.showerror("错误", "模板缺少界面配置")
            return False

        cols = list(self.data.columns)
        enabled_fields = list(state.get("levels", []))
        if state.get("value_enabled", True) and state.get("value_field"):
            enabled_fields.append(state.get("value_field"))
        if state.get("area_enabled", False) and state.get("ratio_field"):
            enabled_fields.append(state.get("ratio_field"))
        if state.get("batch_enabled", False):
            enabled_fields.extend(state.get("batch_fields", []))
        if state.get("group_batch_enabled", False):
            enabled_fields.extend(state.get("group_batch_fields", []))

        missing_fields = sorted({field for field in enabled_fields if field and field not in cols})
        if missing_fields:
            messagebox.showwarning("模板无法应用", "当前数据缺少以下字段：\n" + "\n".join(missing_fields))
            return False

        self.saved_ui_state = state
        self._restore_ui_state(cols)
        stats = [stat for stat in payload.get("stats", []) if stat in self.all_stats]
        if stats:
            self.last_stats = stats
        self.active_group_template_name = template_name
        self._update_loaded_status()
        return True

    def open_group_template_manager(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("分组模板")
        dialog.geometry("720x420")
        dialog.transient(self.root)

        container = ttk.Frame(dialog, padding=12)
        container.pack(fill="both", expand=True)
        container.columnconfigure(1, weight=1)
        container.rowconfigure(0, weight=1)

        left = ttk.Frame(container)
        left.grid(row=0, column=0, sticky="ns", padx=(0, 12))
        right = ttk.Frame(container)
        right.grid(row=0, column=1, sticky="nsew")
        right.rowconfigure(1, weight=1)

        ttk.Label(left, text="已保存模板").pack(anchor="w")
        listbox = tk.Listbox(left, width=24, height=16)
        listbox.pack(fill="y", expand=True, pady=(6, 8))

        ttk.Label(right, text="模板内容").grid(row=0, column=0, sticky="w")
        preview = tk.Text(right, wrap="word", state="disabled", height=16)
        preview.grid(row=1, column=0, sticky="nsew", pady=(6, 8))

        buttons = ttk.Frame(right)
        buttons.grid(row=2, column=0, sticky="e")

        def selected_name():
            selection = listbox.curselection()
            if not selection:
                return ""
            return listbox.get(selection[0])

        def refresh_templates(target_name=""):
            names = sorted(self.group_templates)
            listbox.delete(0, "end")
            for name in names:
                listbox.insert("end", name)

            if target_name and target_name in names:
                index = names.index(target_name)
                listbox.selection_set(index)
                listbox.see(index)
            elif names:
                listbox.selection_set(0)

            update_preview()

        def update_preview(_event=None):
            name = selected_name()
            content = "暂无模板"
            if name:
                content = self._describe_group_template(name, self.group_templates.get(name, {}))
            preview.configure(state="normal")
            preview.delete("1.0", "end")
            preview.insert("1.0", content)
            preview.configure(state="disabled")

        def save_current_template():
            if self.data.empty:
                messagebox.showwarning("提示", "请先加载数据后再保存模板")
                return

            initial_name = self.active_group_template_name or selected_name() or "常用模板"
            name = simpledialog.askstring("保存模板", "请输入模板名称：", parent=dialog, initialvalue=initial_name)
            if not name:
                return
            name = name.strip()
            if not name:
                return

            self.group_templates[name] = self._collect_group_template()
            self.active_group_template_name = name
            self.save_config(show_msg=False)
            refresh_templates(name)
            set_status(self, f"模板“{name}”已保存")

        def apply_selected_template():
            name = selected_name()
            if not name:
                messagebox.showinfo("提示", "请先选择一个模板")
                return
            if self._apply_group_template(self.group_templates.get(name, {}), name):
                self.save_config(show_msg=False)
                dialog.destroy()

        def delete_selected_template():
            name = selected_name()
            if not name:
                return
            if not messagebox.askyesno("确认删除", f"确定要删除模板“{name}”吗？", parent=dialog):
                return
            self.group_templates.pop(name, None)
            if self.active_group_template_name == name:
                self.active_group_template_name = ""
            self.save_config(show_msg=False)
            refresh_templates()

        def export_selected_template():
            name = selected_name()
            if not name:
                messagebox.showinfo("提示", "请先选择一个模板")
                return
            path = filedialog.asksaveasfilename(
                title="导出模板",
                defaultextension=".json",
                filetypes=[("JSON 文件", "*.json"), ("所有文件", "*.*")],
                initialfile=f"{name}.json",
            )
            if not path:
                return
            with open(path, "w", encoding="utf-8") as file:
                json.dump({"name": name, "template": self.group_templates[name]}, file, ensure_ascii=False, indent=2)
            set_status(self, f"模板已导出到 {os.path.basename(path)}")

        def import_template():
            path = filedialog.askopenfilename(
                title="导入模板",
                filetypes=[("JSON 文件", "*.json"), ("所有文件", "*.*")],
            )
            if not path:
                return

            with open(path, "r", encoding="utf-8") as file:
                payload = json.load(file)

            if isinstance(payload, dict) and "template" in payload:
                name = str(payload.get("name") or os.path.splitext(os.path.basename(path))[0]).strip()
                template = payload.get("template", {})
            else:
                name = os.path.splitext(os.path.basename(path))[0]
                template = payload

            if not name or not isinstance(template, dict):
                messagebox.showerror("导入失败", "模板文件格式不正确")
                return

            self.group_templates[name] = template
            self.save_config(show_msg=False)
            refresh_templates(name)
            set_status(self, f"模板“{name}”已导入")

        ttk.Button(buttons, text="保存当前", command=save_current_template).pack(side="left", padx=(0, 6))
        ttk.Button(buttons, text="应用", command=apply_selected_template).pack(side="left", padx=6)
        ttk.Button(buttons, text="删除", command=delete_selected_template).pack(side="left", padx=6)
        ttk.Button(buttons, text="导出", command=export_selected_template).pack(side="left", padx=6)
        ttk.Button(buttons, text="导入", command=import_template).pack(side="left", padx=6)
        ttk.Button(buttons, text="关闭", command=dialog.destroy).pack(side="left", padx=(6, 0))

        listbox.bind("<<ListboxSelect>>", update_preview)
        refresh_templates(self.active_group_template_name)
        dialog.grab_set()
        self.root.wait_window(dialog)

    def open_filter_builder(self):
        if self.data.empty:
            messagebox.showwarning("提示", "请先加载数据后再设置筛选条件")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("条件筛选")
        dialog.geometry("760x420")
        dialog.transient(self.root)

        container = ttk.Frame(dialog, padding=12)
        container.pack(fill="both", expand=True)
        container.rowconfigure(1, weight=1)
        container.columnconfigure(0, weight=1)

        ttk.Label(container, text=f"当前数据共 {len(self.data)} 行，可添加多个筛选条件。").grid(row=0, column=0, sticky="w")

        rows_frame = ttk.Frame(container)
        rows_frame.grid(row=1, column=0, sticky="nsew", pady=(10, 8))
        rows_frame.columnconfigure(0, weight=1)

        summary_var = tk.StringVar()
        detail_var = tk.StringVar(value=self._filter_summary())
        condition_rows = []
        cols = list(self.data.columns)

        def collect_conditions():
            conditions = []
            for item in condition_rows:
                field = item["field"].get().strip()
                op = item["op"].get().strip()
                value = item["value"].get().strip()
                enabled = bool(item["enabled"].get())
                if field and op:
                    conditions.append(
                        {
                            "field": field,
                            "op": op,
                            "value": value,
                            "enabled": enabled,
                        }
                    )
            return conditions

        def refresh_summary():
            conditions = collect_conditions()
            active_conditions = [c for c in conditions if c.get("enabled", True)]
            detail_var.set(describe_filter_conditions(active_conditions))
            try:
                filtered = apply_filter_conditions(self.data, conditions)
            except Exception as exc:
                summary_var.set(f"条件有误：{exc}")
                return

            if active_conditions:
                summary_var.set(f"已启用 {len(active_conditions)} 条，筛选后 {len(filtered)} / {len(self.data)} 行")
            else:
                summary_var.set(f"当前未启用筛选，将使用全部 {len(self.data)} 行数据")

        def remove_condition(item):
            item["frame"].destroy()
            if item in condition_rows:
                condition_rows.remove(item)
            if not condition_rows:
                add_condition_row()
            refresh_summary()

        def add_condition_row(condition=None):
            condition = condition or {}
            frame = ttk.Frame(rows_frame)
            frame.pack(fill="x", pady=4)

            enabled_var = tk.BooleanVar(value=bool(condition.get("enabled", True)))
            field_var = tk.StringVar(value=str(condition.get("field", cols[0] if cols else "")))
            op_var = tk.StringVar(value=str(condition.get("op", FILTER_OPERATORS[0])))
            value_var = tk.StringVar(value=str(condition.get("value", "")))

            ttk.Checkbutton(frame, text="启用", variable=enabled_var, command=refresh_summary).pack(side="left")
            field_box = ttk.Combobox(frame, values=cols, textvariable=field_var, state="readonly", width=20)
            field_box.pack(side="left", padx=(8, 6))
            op_box = ttk.Combobox(frame, values=FILTER_OPERATORS, textvariable=op_var, state="readonly", width=12)
            op_box.pack(side="left", padx=6)
            value_entry = ttk.Entry(frame, textvariable=value_var, width=22)
            value_entry.pack(side="left", padx=6)

            item = {
                "frame": frame,
                "enabled": enabled_var,
                "field": field_var,
                "op": op_var,
                "value": value_var,
            }

            def sync_value_state(*_args):
                disabled = op_var.get() in {"为空", "不为空"}
                value_entry.configure(state="disabled" if disabled else "normal")
                if disabled:
                    value_var.set("")
                refresh_summary()

            op_var.trace_add("write", sync_value_state)
            field_box.bind("<<ComboboxSelected>>", lambda _event: refresh_summary())
            value_entry.bind("<KeyRelease>", lambda _event: refresh_summary())
            ttk.Button(frame, text="删除", command=lambda: remove_condition(item)).pack(side="left", padx=(6, 0))
            condition_rows.append(item)
            sync_value_state()

        controls = ttk.Frame(container)
        controls.grid(row=2, column=0, sticky="ew")
        ttk.Button(controls, text="新增条件", command=add_condition_row).pack(side="left")
        ttk.Button(
            controls,
            text="清空筛选",
            command=lambda: (
                [row["frame"].destroy() for row in condition_rows],
                condition_rows.clear(),
                self.filter_conditions.clear(),
                add_condition_row(),
                refresh_summary(),
            ),
        ).pack(side="left", padx=6)

        ttk.Label(container, textvariable=summary_var).grid(row=3, column=0, sticky="w", pady=(10, 2))
        ttk.Label(container, textvariable=detail_var, wraplength=700).grid(row=4, column=0, sticky="w")

        buttons = ttk.Frame(container)
        buttons.grid(row=5, column=0, sticky="e", pady=(12, 0))

        def apply_conditions():
            conditions = collect_conditions()
            try:
                filtered = apply_filter_conditions(self.data, conditions)
            except Exception as exc:
                messagebox.showerror("筛选条件有误", str(exc), parent=dialog)
                return

            self.filter_conditions = [c for c in conditions if c.get("enabled", True)]
            self._update_loaded_status()
            self.save_config(show_msg=False)
            if self.filter_conditions:
                set_status(self, f"筛选已更新，当前保留 {len(filtered)} 行")
            else:
                set_status(self, f"已清除筛选，当前共 {len(self.data)} 行")
            dialog.destroy()

        ttk.Button(buttons, text="应用", command=apply_conditions).pack(side="left", padx=(0, 6))
        ttk.Button(buttons, text="关闭", command=dialog.destroy).pack(side="left")

        existing_conditions = self.filter_conditions or [{}]
        for condition in existing_conditions:
            add_condition_row(condition)
        refresh_summary()
        dialog.grab_set()
        self.root.wait_window(dialog)

    def open_stats_order(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("自定义统计量顺序")
        dlg.geometry("300x400")

        ttk.Label(dlg, text="拖拽或用按钮调整顺序：").pack(pady=6)

        lb = tk.Listbox(dlg)
        lb.pack(fill='both', expand=True, padx=12, pady=6)
        for stat in self.last_stats:
            lb.insert('end', stat)

        # 拖拽排序支持
        def _on_start(e):
            lb._drag_index = lb.nearest(e.y)

        def _on_motion(e):
            i = lb.nearest(e.y)
            if i != lb._drag_index:
                v = lb.get(lb._drag_index)
                lb.delete(lb._drag_index)
                lb.insert(i, v)
                lb.select_clear(0, 'end')
                lb.select_set(i)
                lb._drag_index = i

        lb.bind("<Button-1>", _on_start)
        lb.bind("<B1-Motion>", _on_motion)

        # 上下移动按钮
        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(pady=4)


        def move_up():
            sel = lb.curselection()
            if not sel or sel[0] == 0: return
            i = sel[0];
            v = lb.get(i)
            lb.delete(i);
            lb.insert(i - 1, v);
            lb.select_set(i - 1)

        def move_down():
            sel = lb.curselection()
            if not sel or sel[0] == lb.size() - 1: return
            i = sel[0];
            v = lb.get(i)
            lb.delete(i);
            lb.insert(i + 1, v);
            lb.select_set(i + 1)

        ttk.Button(btn_frame, text="上移", command=move_up).grid(row=0, column=0, padx=6)
        ttk.Button(btn_frame, text="下移", command=move_down).grid(row=0, column=1, padx=6)

        # 确定后更新 self.last_stats
        def on_ok():
            self.last_stats = [lb.get(i) for i in range(lb.size())]
            dlg.destroy()

        ttk.Button(dlg, text="确定", command=on_ok, width=10) \
            .pack(pady=(0, 12))

        dlg.grab_set()
        self.root.wait_window(dlg)

    def open_custom_sort(self):
        # 先收集所有已选的分类级别列
        cols = [l["var"].get() for l in self.levels if l["var"].get()]

        # 如果批量分组开启，把它们也加入可选列
        if getattr(self, "group_batch_fields", None):
            for g in self.group_batch_fields:
                if g not in cols:
                    cols.append(g)

        if not cols:
            messagebox.showwarning("提示", "请先设置分类级别")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("自定义排序")
        dlg.geometry("300x440")

        ttk.Label(dlg, text="请选择列:").pack(pady=(10,0))
        col_var = tk.StringVar(value=cols[0])
        col_cmb = ttk.Combobox(dlg, values=cols, textvariable=col_var,
                               state="readonly")
        col_cmb.pack(fill='x', padx=12, pady=6)

        lb = tk.Listbox(dlg)
        lb.pack(fill='both', expand=True, padx=12, pady=6)
        last_col = None
        # 拖拽排序
        def _on_start_drag(e):
            lb._drag_index = lb.nearest(e.y)
        def _on_drag_motion(e):
            i = lb.nearest(e.y)
            if i != lb._drag_index:
                v = lb.get(lb._drag_index)
                lb.delete(lb._drag_index)
                lb.insert(i, v)
                lb.select_clear(0, 'end')
                lb.select_set(i)
                lb._drag_index = i

        lb.bind("<Button-1>", _on_start_drag)
        lb.bind("<B1-Motion>", _on_drag_motion)

        # 上/下移
        btnf = ttk.Frame(dlg); btnf.pack(pady=4)
        def up():
            sel = lb.curselection()
            if not sel or sel[0]==0: return
            i=sel[0];v=lb.get(i)
            lb.delete(i); lb.insert(i-1,v); lb.select_set(i-1)
        def down():
            sel = lb.curselection()
            if not sel or sel[0]==lb.size()-1: return
            i=sel[0];v=lb.get(i)
            lb.delete(i); lb.insert(i+1,v); lb.select_set(i+1)
        ttk.Button(btnf, text="上移",   command=up).grid(row=0,column=0,padx=6)
        ttk.Button(btnf, text="下移",   command=down).grid(row=0,column=1,padx=6)

        # 加载列表初始值
        def load_vals(e=None):
            nonlocal last_col
            if last_col:
                self.custom_orders[last_col] = list(lb.get(0, 'end'))
            col = col_var.get()
            # 记录原始顺序（只第一次记录）
            if col not in self.original_orders:
                vals = list(self.data[col].dropna().unique())
                self.original_orders[col] = vals.copy()
            # 先清空
            lb.delete(0, 'end')

            seq = self.custom_orders.get(col, self.original_orders[col])
            for v in seq:
                lb.insert('end', v)
                # 切换完成后，将本次列名记为 last_col
            last_col = col

        load_vals()
        col_cmb.bind("<<ComboboxSelected>>", load_vals)

        # 导入 TXT（保持你已有逻辑）
        def import_txt():
            path = filedialog.askopenfilename(
                title="导入排序 TXT",
                filetypes=[("Text 文件","*.txt"),("All","*.*")]
            )
            if not path: return
            with open(path, encoding='utf-8') as f:
                lines = [l.strip() for l in f if l.strip()]
            curr = list(lb.get(0,'end'))
            if len(lines)!=len(curr) or set(lines)!=set(curr):
                cont = messagebox.askyesno(
                    "导入警告",
                    f"TXT 行数 {len(lines)} 与当前唯一值 {len(curr)} 不匹配，是否继续导入？"
                )
                if not cont: return
            lb.delete(0,'end')
            for v in lines: lb.insert('end', v)

        # 恢复原始顺序
        def restore_original():
            col = col_var.get()
            if col in self.original_orders:
                lb.delete(0,'end')
                for v in self.original_orders[col]:
                    lb.insert('end', v)

        # 确定
        def on_ok():
            col = col_var.get()
            raw = list(lb.get(0, 'end'))
            # —— 如果原始列是数值型，就把字符串转成对应数字
            if pd.api.types.is_numeric_dtype(self.data[col]):
                dtype = self.data[col].dtype
                def to_num(x):
                    try:
                        # 整数列用 int，浮点列用 float
                        return int(x) if np.issubdtype(dtype, np.integer) else float(x)
                    except:
                        return float(x)
                self.custom_orders[col] = [to_num(v) for v in raw]
            else:
                # 非数值型，保留字符串排序
                self.custom_orders[col] = raw
            dlg.destroy()

        btn_tools = ttk.Frame(dlg); btn_tools.pack(pady=(0,8))
        ttk.Button(btn_tools, text="从 TXT 导入", command=import_txt)\
            .grid(row=0, column=0, padx=6)
        ttk.Button(btn_tools, text="恢复原始顺序",
                   command=restore_original)\
            .grid(row=0, column=1, padx=6)

        # 定义导出函数
        def export_txt():
            path = filedialog.asksaveasfilename(
                title="导出排序 TXT",
                defaultextension=".txt",
                filetypes=[("Text 文件", "*.txt"), ("All", "*.*")]
            )
            if not path:
                return

            # 读取 Listbox 中的当前顺序
            lines = [lb.get(i) for i in range(lb.size())]

            try:
                with open(path, 'w', encoding='utf-8') as f:
                    for v in lines:
                        f.write(f"{v}\n")
                messagebox.showinfo("导出成功", f"已将当前排序导出到：\n{path}")
            except Exception as e:
                messagebox.showerror("导出失败", str(e))

        # 在工具按钮区添加 “导出排序” 按钮
        ttk.Button(btn_tools, text="导出排序", command=export_txt) \
            .grid(row=0, column=2, padx=6)

        ttk.Button(dlg, text="确定", command=on_ok, width=10)\
            .pack(pady=(0,12))

        dlg.grab_set()
        self.root.wait_window(dlg)

    def toggle_batch(self):
        if self.batch_var.get():
            self.batch_btn.state(["!disabled"])
        else:
            self.batch_btn.state(["disabled"])
            self.batch_fields = []

    def choose_batch_fields(self):
        if self.data.empty:
            messagebox.showwarning("提示", "请先加载数据")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("选择值字段(多选)")
        cols = list(self.data.columns)
        lb = tk.Listbox(dlg, selectmode="multiple", height=10)
        for c in cols:
            lb.insert("end", c)
        lb.pack(padx=12, pady=6)
        def on_ok():
            self.batch_fields = [cols[i] for i in lb.curselection()]
            dlg.destroy()
        ttk.Button(dlg, text="确定", command=on_ok, width=10).pack(pady=6)
        dlg.grab_set()
        self.root.wait_window(dlg)

    # 批量分组字段(分类级别1)
    def toggle_group_batch(self):
        if self.group_batch_var.get():
            self.group_batch_btn.state(["!disabled"])
        else:
            self.group_batch_btn.state(["disabled"])
            self.group_batch_fields = []

    def choose_group_batch_fields(self):
        if self.data.empty:
            messagebox.showwarning("提示", "请先加载数据")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("选择分组字段(多选) —— 作为 分类级别1")
        cols = list(self.data.columns)
        lb = tk.Listbox(dlg, selectmode="multiple", height=10)
        for c in cols:
            lb.insert("end", c)
        lb.pack(padx=12, pady=6)
        def on_ok():
            self.group_batch_fields = [cols[i] for i in lb.curselection()]
            dlg.destroy()
        ttk.Button(dlg, text="确定", command=on_ok, width=10).pack(pady=6)
        dlg.grab_set()
        self.root.wait_window(dlg)

    # 整表总计行
    def _append_overall_total(self, final, df, gf, val, ratio, need_area):
        stats_map = calculate_series_stats(df[val], int(pd.to_numeric(df[val], errors="coerce").count()))
        stats_map["数量占比"] = round2(100.0)
        if need_area:
            total_area = df[ratio].sum()

            dec = self.area_decimals_var.get()
            fmt = "0" if dec == 0 else "0." + "0" * dec
            stats_map["面积(亩)"] = float(Decimal(str(total_area))
                                          .quantize(Decimal(fmt), ROUND_HALF_UP))
            stats_map["面积占比"] = round2(100.0)

        total_row = {}
        for col in final.columns:
            if col in gf:
                total_row[col] = "总计" if col == gf[0] else ""
            else:
                total_row[col] = stats_map.get(col, "")

        overall_df = pd.DataFrame([total_row], columns=final.columns)
        return pd.concat([final, overall_df], ignore_index=True)

    def _round_result_frame(self, df, group_fields, need_area):
        dec = self.area_decimals_var.get()

        def round_dec(value, digits):
            try:
                fmt = "0" if digits == 0 else "0." + "0" * digits
                return float(Decimal(str(value)).quantize(Decimal(fmt), ROUND_HALF_UP))
            except Exception:
                return value

        for column in df.columns:
            if column in group_fields or column in {"数量", "缺失数"}:
                continue
            if column == "面积(亩)" and need_area:
                df[column] = df[column].apply(lambda value: round_dec(value, dec))
            else:
                df[column] = df[column].apply(round2)
        return df

    def _write_export_overview(self, writer, output_path, overview_rows):
        active_filters = [c for c in self.filter_conditions if c.get("enabled", True)]
        filtered_rows = len(self.get_active_data()) if hasattr(self, "get_active_data") else len(self.data)
        summary_rows = [
            {"项目": "导出时间", "内容": time.strftime("%Y-%m-%d %H:%M:%S")},
            {"项目": "模式", "内容": self.mode_label},
            {"项目": "源文件", "内容": self.excel_path},
            {"项目": "数据源", "内容": self.current_sheet_name or "当前数据"},
            {"项目": "原始记录数", "内容": len(self.data)},
            {"项目": "导出记录数", "内容": filtered_rows},
            {"项目": "导出文件", "内容": output_path},
            {"项目": "统计量", "内容": "、".join(self.last_stats)},
            {"项目": "插入小计行", "内容": "是" if self.subtotal_var.get() else "否"},
            {"项目": "面积统计", "内容": "是" if self.area_cb.get() else "否"},
            {"项目": "筛选条件", "内容": describe_filter_conditions(active_filters)},
            {"项目": "当前模板", "内容": self.active_group_template_name or "未使用"},
            {"项目": "批量值字段", "内容": "、".join(self.batch_fields) if self.batch_fields else "未启用"},
            {"项目": "批量分组字段", "内容": "、".join(self.group_batch_fields) if self.group_batch_fields else "未启用"},
        ]
        pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name="统计说明")
        if overview_rows:
            pd.DataFrame(overview_rows).to_excel(writer, index=False, sheet_name="导出清单")

    # 文件 & 数据加载
    def toggle_area_fields(self):
        if self.area_cb.get():
            self.ratio_menu.state(["!disabled"])
            self.area_decimals_spin.state(["!disabled"])
        else:
            self.ratio_menu.state(["disabled"])
            self.ratio_var.set("")
            self.area_decimals_spin.state(["disabled"])

    def toggle_value_field(self):
        """控制值字段下拉和批量值字段复选框的启用/禁用"""
        if self.val_enable_var.get():
                self.val_menu.state(["!disabled"])
                # 保持 batch 的 checkbox 可操作
        else:
                # 禁用下拉，并清空已选
                self.val_menu.state(["disabled"])
                self.val_var.set("")

                self.batch_var.set(False)
                self.toggle_batch()
    def select_file(self):
        self._show_progress("选择文件中...")
        path = filedialog.askopenfilename(
            filetypes=[
                ("Excel 文件", ("*.xlsx", "*.xls")),
                ("CSV 文件", "*.csv"),
                ("所有文件", "*.*")
            ]
        )
        if not path:
            return self._hide_progress("就绪")
        self.add_recent(path)
        self.excel_path = path
        self.lbl_file.config(text=os.path.basename(path))
        ext = os.path.splitext(path)[1].lower()

        if ext == ".csv":
            threading.Thread(target=self._read_csv, daemon=True).start()
        else:
            threading.Thread(target=self._get_sheets, daemon=True).start()

    def handle_drop(self, e):
        path = e.data.strip("{}")
        ext = os.path.splitext(path)[1].lower()
        if ext not in (".xlsx", ".xls", ".csv"):
            return messagebox.showerror("错误", "只支持拖拽 .xlsx、.xls 或 .csv")
        self._show_progress("读取文件中...")
        self.add_recent(path)
        self.excel_path = path
        self.lbl_file.config(text=os.path.basename(path))

        if ext == ".csv":
            threading.Thread(target=self._read_csv, daemon=True).start()
        else:
            threading.Thread(target=self._get_sheets, daemon=True).start()

    def _get_sheets(self):
        try:
            sheets = pd.ExcelFile(self.excel_path).sheet_names
        except Exception as ex:
            self.root.after(0, lambda: messagebox.showerror("错误", f"{ex}"))
            return self._hide_progress("就绪")
        self.root.after(0, lambda: self._on_sheets(sheets))

    def _on_sheets(self, sheets):
        self._hide_progress()
        self.sheet_menu["values"] = sheets
        if sheets:
            self.sheet_var.set(sheets[0])
            self.load_sheet(sheets[0])

    def load_sheet(self, sheet):
        self.calculate_btn.state(["disabled"])
        self._show_progress(f"加载子表: {sheet}")
        threading.Thread(target=self._read_sheet, args=(sheet,), daemon=True).start()

    def _read_sheet(self, sheet):
        try:
            df = pd.read_excel(self.excel_path, sheet_name=sheet)
        except Exception as ex:
            self.root.after(0, lambda: messagebox.showerror("错误", f"{ex}"))
            return self._hide_progress("就绪")
        self.root.after(0, lambda: self._on_data(df))

    def _read_csv(self):
        try:
            df = read_csv_safely(self.excel_path)
        except Exception as ex:
            self.root.after(0, lambda: messagebox.showerror("错误", str(ex)))
            return self._hide_progress("就绪")
        self.root.after(0, lambda: self._on_data(df))

    def _on_data(self, df):
        self._hide_progress()
        self.data = df
        ext = os.path.splitext(self.excel_path)[1].lower()
        self.current_sheet_name = "CSV" if ext == ".csv" else (self.sheet_var.get() or "当前子表")

        cols = list(df.columns)
        for cmb in (self.val_menu, self.ratio_menu):
            cmb["values"] = cols
        if cols:
            self.val_var.set(cols[-1])
            self.ratio_var.set(cols[-1])
        self._refresh_levels(cols)
        self._restore_ui_state(cols)
        self._normalize_filter_conditions(cols)

        if ext == ".csv":
            self.sheet_menu.set("")
            self.sheet_menu["values"] = []
            self.sheet_menu.state(["disabled"])
        else:
            if "disabled" in self.sheet_menu.state():
                self.sheet_menu.state(["!disabled"])

        self.toggle_area_fields()
        self.calculate_btn.state(["!disabled"])
        self._update_loaded_status()

    def _refresh_levels(self, cols):
        for lvl in self.levels:
            c = lvl["combo"]
            c["values"] = cols
            if not c.get() or c.get() not in cols:
                c.set(cols[0] if cols else "")

    def add_level(self):
        if len(self.levels) >= MAX_LEVELS:
            return
        var = tk.StringVar()
        lbl = ttk.Label(self.frame_levels, text=f"分类级 {len(self.levels)+1}:")
        cmb = ttk.Combobox(
            self.frame_levels, textvariable=var,
            state="readonly", width=22
        )
        btn = ttk.Button(self.frame_levels, text="×", width=2)
        btn.config(command=lambda b=btn: self._remove_level(b))
        self.levels.append({"var":var, "label":lbl, "combo":cmb, "btn":btn})
        if not self.data.empty:
            cs = list(self.data.columns)
            cmb["values"] = cs
            cmb.set(cs[0])
        self._layout()

    def _remove_level(self, btn):
        for i, l in enumerate(self.levels):
            if l["btn"] is btn:
                l["label"].destroy()
                l["combo"].destroy()
                l["btn"].destroy()
                self.levels.pop(i)
                break
        self._layout()

    def _layout(self):
        for i, l in enumerate(self.levels):
            l["label"].grid(row=i, column=0, padx=6, pady=2, sticky='w')
            l["combo"].grid(row=i, column=1, padx=6, pady=2, sticky='w')
            l["btn"].grid(row=i, column=2, padx=6)
        self.add_btn.grid(row=len(self.levels), column=0, columnspan=3, pady=6)

    def choose_stats(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("选择统计量")
        dlg.geometry("560x360")
        frm = ttk.Frame(dlg)
        frm.pack(padx=12, pady=12)
        ttk.Label(frm, text="请选择统计量:").pack(pady=(0,8))
        chkf = ttk.Frame(frm)
        chkf.pack()
        # 用两个 dict 分别保存 BooleanVar 和对应的 Checkbutton
        tmp = {}  # {stat_name: BooleanVar}
        cbs = {}  # {stat_name: Checkbutton}
        for i, st in enumerate(self.all_stats):
            v = tk.BooleanVar(value=(st in self.last_stats))
            cb = ttk.Checkbutton(chkf, text=st, variable=v)
            cb.grid(row=i//4, column=i%4, padx=8, pady=5, sticky='w')
            tmp[st] = v
            cbs[st] = cb
        ttk.Button(frm, text="确定", command=dlg.destroy, width=10).pack(pady=8)
        dlg.grab_set()
        self.root.wait_window(dlg)
        self.last_stats = [s for s, v in tmp.items() if v.get()]

    def calculate(self):
        if not self.val_enable_var.get() and not self.area_cb.get() and not self.batch_var.get():
            messagebox.showwarning("提示",
                                    "请勾选“启用值字段”、“计算面积占比”或“批量值字段”后再导出"
            )
            return
        try:
            self._calculate()
        except Exception as ex:
            self.status_var.set(f"计算错误: {ex}")
            messagebox.showerror("错误", traceback.format_exc())


    def create_recent_menu(self):
        """根据 self.recent_files 构建最近打开菜单。"""
        populate_recent_menus(self)

    def add_recent(self, path):
        """将 path 插入到最近打开列表顶端，去重并限长20，静默保存。"""
        if path in self.recent_files:
            self.recent_files.remove(path)
        self.recent_files.insert(0, path)
        self.recent_files = self.recent_files[:20]
        self.save_config(show_msg=False)
        self.create_recent_menu()

    def open_recent(self, path):
        """通过菜单快速打开已有文件。"""
        if not os.path.exists(path):
            messagebox.showerror("文件不存在", path)
            self.recent_files.remove(path)
            self.save_config(show_msg=False)
            self.create_recent_menu()
            return

        self.excel_path = path
        self.lbl_file.config(text=os.path.basename(path))

        ext = os.path.splitext(path)[1].lower()
        if ext == ".csv":
            threading.Thread(target=self._read_csv, daemon=True).start()
        else:
            threading.Thread(target=self._get_sheets, daemon=True).start()

        # 更新最近打开顺序
        self.add_recent(path)

    def _calculate(self):
        start = time.time()
        self._show_progress("正在计算...")
        if self.data.empty:
            messagebox.showerror("错误", "请先加载数据")
            return self._hide_progress("就绪")

        active_df = self._get_filtered_data(show_error=True)
        if active_df is None:
            return self._hide_progress("就绪")
        if active_df.empty:
            messagebox.showwarning("提示", "筛选后没有可用数据，请调整条件后再试。")
            return self._hide_progress("就绪")

        # 默认分类级别
        gf = [l["var"].get() for l in self.levels if l["var"].get()]
        if self.batch_var.get() and getattr(self, "batch_fields", None):
            fields = self.batch_fields or [self.val_var.get()]
        else:
            fields = [self.val_var.get()]

        # 批量分组字段
        groups = []
        if self.group_batch_var.get() and getattr(self, "group_batch_fields", None):
            groups = self.group_batch_fields

        # 先取到当前是否开启批量值字段 / 批量分组
        batch_on = self.batch_var.get()
        group_on = self.group_batch_var.get()

        # 取出 fields, groups
        if batch_on and getattr(self, "batch_fields", None):
            fields = self.batch_fields or fields

        groups = []
        if group_on and getattr(self, "group_batch_fields", None):
            groups = self.group_batch_fields or []

        # --------- 这里替换原来的 loops 逻辑 ------------
        loops = []
        if batch_on and group_on:
            # 如果一一对应
            if len(fields) == len(groups):
                loops = [(f, [g]) for f, g in zip(fields, groups)]
            else:
                # 弹窗确认
                cont = messagebox.askyesno(
                    "提示",
                    "所选值字段和分组字段数量不一致，\n"
                    "是否按“所有值字段 × 所有分组字段”计算？"
                )
                if not cont:
                    # 用户点“否”，就直接退出
                    return self._hide_progress("就绪")
                # 真正的笛卡尔积：所有 (f,g) 对
                loops = [(f, [g]) for f, g in itertools.product(fields, groups)]

        elif batch_on:
            # 只有批量值字段
            loops = [(f, self.levels and [l["var"].get() for l in self.levels if l["var"].get()] or []) for f in fields]

        elif group_on:
            # 只有批量分组字段
            loops = [(self.val_var.get(), [g]) for g in groups]

        else:
            # 都没批量
            loops = [(self.val_var.get(),
                      [l["var"].get() for l in self.levels if l["var"].get()])]

        ratio = self.ratio_var.get()
        need_area = self.area_cb.get()
        need_sub = bool(self.subtotal_var.get())

        basename = os.path.splitext(os.path.basename(self.excel_path))[0]
        out = build_output_path(self, f"{basename}_结果")

        # 如果没有任何可统计的(field, group)组合，提前退出
        if not loops:
            messagebox.showerror("错误", "没有可统计的值/分组组合，无法导出")
            return self._hide_progress("就绪")

        overview_rows = []
        skipped = []
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            for field, gf_current in loops:
                self.status_var.set(f"计算: 值字段={field}, 分组={gf_current}")

                df0 = active_df.copy()
                df0[field] = pd.to_numeric(df0[field], errors="coerce")
                valid_count = int(df0[field].count())
                if valid_count == 0:
                    skipped.append((field, gf_current))
                    continue

                # 仅在需要面积占比时才处理 ratio 列
                if need_area:
                    df0[ratio] = pd.to_numeric(df0[ratio], errors="coerce")
                    area_total = df0[ratio].sum()
                else:
                    area_total = None

                res = build_grouped_stats_frame(df0, field, gf_current)

                if need_area:
                    if gf_current:
                        area_df = df0.groupby(gf_current, dropna=False)[ratio].sum().reset_index(name="面积(亩)")
                        res = res.merge(area_df, on=gf_current, how="left")
                        total_area = res["面积(亩)"].sum()
                    else:
                        total_area = float(area_total) if area_total is not None else 0.0
                        res["面积(亩)"] = total_area
                    res["面积占比"] = (res["面积(亩)"] / total_area * 100) if total_area else np.nan

                res = self._round_result_frame(res, gf_current, need_area)
                filtered = apply_count_masks(res)

                # 列切片
                display_cols = gf_current + self.last_stats.copy()
                if need_area:
                    display_cols += ["面积(亩)", "面积占比"]
                tmp = filtered[display_cols].copy()

                # —— 多级稳定排序 ——
                if gf_current:
                    for col in reversed(gf_current):
                        if col in self.custom_orders:
                            # 只保留本轮实际出现的顺序 —— 先转成 Python list 再判断
                            present_vals = tmp[col].dropna().unique().tolist()
                            uniq = [v for v in self.custom_orders[col]
                                    if v in present_vals]
                            tmp[col] = pd.Categorical(
                                tmp[col],
                                categories=uniq,
                                ordered=True
                            )
                            tmp = tmp.sort_values(by=col, kind="stable")
                        else:
                            tmp = tmp.sort_values(
                                by=col,
                                key=lambda s: s.map(sort_key),
                                kind="stable"
                            )

                # 小计逻辑：仅当 need_sub 且有分组时，并且该组行数 >1 才输出小计
                if need_sub and gf_current:
                    def _group_key(value):
                        return "__NA__" if pd.isna(value) else value

                    grp_counts = {_group_key(key): count for key, count in tmp[gf_current[0]].value_counts(dropna=False).items()}

                    out_rows = []
                    prev = None
                    prev_key = None
                    for _, r in tmp.iterrows():
                        cur = r[gf_current[0]]
                        cur_key = _group_key(cur)
                        # 切组边界时，只有上一个组大小>1，才插小计
                        if prev is not None and cur_key != prev_key and grp_counts.get(prev_key, 0) > 1:
                            out_rows.append(
                                self._subtotal(
                                    prev, df0, gf_current, field, ratio, need_area
                                )
                            )
                        out_rows.append(r.to_dict())
                        prev = cur
                        prev_key = cur_key
                    # 最后一组结束后，若该组大小>1，则插小计
                    if prev is not None and grp_counts.get(prev_key, 0) > 1:
                        out_rows.append(
                            self._subtotal(
                                prev, df0, gf_current, field, ratio, need_area
                            )
                        )
                    final_df = pd.DataFrame(out_rows, columns=tmp.columns)
                else:
                    final_df = tmp.copy()

                # 整表总计行
                final_df = self._append_overall_total(
                    final_df, df0, gf_current, field,
                    ratio, need_area
                )

                # 写入 Sheet
                name_parts = [field] + gf_current
                sheet_name = "_".join(name_parts)[:31]
                final_df.to_excel(writer, index=False, sheet_name=sheet_name)
                ws = writer.sheets[sheet_name]
                overview_rows.append({
                    "工作表": sheet_name,
                    "值字段": field,
                    "分组字段": " / ".join(gf_current) if gf_current else "无分组",
                    "导出行数": len(final_df),
                    "有效数量": valid_count,
                    "面积统计": "是" if need_area else "否",
                    "小计行": "是" if need_sub and gf_current else "否",
                })

                thin = Side(border_style='thin', color='000000')
                bd = Border(thin, thin, thin, thin)

                if gf_current:
                    start_row = 2
                    row_count = len(final_df)

                    # 1) 垂直合并：每一级分类列不跨“小计” nor “总计”
                    for ci in range(len(gf_current)):
                        merge_start = start_row
                        prev = ws.cell(merge_start, ci + 1).value

                        for r in range(start_row + 1, start_row + row_count):
                            curr = ws.cell(r, ci + 1).value

                            # 标记：如果 prev 是小计或总计，就一定切断
                            is_sub = isinstance(prev, str) and prev.endswith(" 合计")
                            is_total = (prev == "总计")

                            # 值变了 或者 刚好是小计/总计，都要切断上一段
                            if curr != prev or is_sub or is_total:
                                # 如果上一段有多行，并且上一段不是合计，就合并
                                if r - merge_start > 1 and not (is_sub or is_total):
                                    ws.merge_cells(
                                        start_row=merge_start,
                                        start_column=ci + 1,
                                        end_row=r - 1,
                                        end_column=ci + 1
                                    )
                                merge_start = r
                                prev = curr

                        # 合并最后一段，排除合计
                        last_span = (start_row + row_count - 1) - merge_start + 1
                        if last_span > 1:
                            if not (isinstance(prev, str) and prev.endswith(" 合计")) \
                                    and prev != "总计":
                                ws.merge_cells(
                                    start_row=merge_start,
                                    start_column=ci + 1,
                                    end_row=start_row + row_count - 1,
                                    end_column=ci + 1
                                )

                    # 2) 横向合并：所有“XXX 合计” 和 最后的“总计”
                    for r in range(start_row, start_row + row_count):
                        v = ws.cell(r, 1).value
                        if (isinstance(v, str) and v.endswith(" 合计")) or v == "总计":
                            ws.merge_cells(
                                start_row=r,
                                start_column=1,
                                end_row=r,
                                end_column=len(gf_current)
                            )

                # 3) 整表样式 & 数字格式
                # 3) 整表样式 & 数字格式 —— 按列名动态应用格式
                dec = self.area_decimals_var.get()
                area_fmt = "0" if dec == 0 else "0." + "0" * dec
                percent_fmt = "0.00"

                # 把 final_df.columns 映射到 Excel 列号
                col_idx_to_name = {i + 1: name for i, name in enumerate(final_df.columns)}

                for row in ws.iter_rows(
                        min_row=1, max_row=ws.max_row,
                        min_col=1, max_col=ws.max_column
                ):
                    for col_idx, cell in enumerate(row, start=1):
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = bd if "bd" in locals() else Border()

                        # 跳过第一行表头
                        if cell.row == 1:
                            continue

                        col_name = col_idx_to_name.get(col_idx, "")
                        if col_name == "面积(亩)" and isinstance(cell.value, (int, float)):
                            # 面积列按用户设置小数位
                            cell.number_format = area_fmt
                        elif col_name == "面积占比" and isinstance(cell.value, (int, float)):
                            # 面积占比固定两位
                            cell.number_format = percent_fmt
                        else:
                            # 其它列：整数 0，小数两位
                            if isinstance(cell.value, int):
                                cell.number_format = "0"
                            elif isinstance(cell.value, float):
                                cell.number_format = "0.00"

            self._write_export_overview(writer, out, overview_rows)

        if not overview_rows:
            if os.path.exists(out):
                os.remove(out)
            messagebox.showerror("错误", "没有找到可用于统计的数值数据。")
            return self._hide_progress("就绪")

        if skipped:
            skipped_text = "\n".join(
                f"- 值字段 {field} / 分组 {' / '.join(groups) if groups else '无分组'}"
                for field, groups in skipped[:8]
            )
            messagebox.showwarning("部分组合已跳过", f"以下组合没有有效数值，已自动跳过：\n{skipped_text}")

        elapsed = round2(time.time() - start)
        self._hide_progress(f"完成:{elapsed}s")
        self.show_result_dialog(out, elapsed)


    def _subtotal(self, fl, df_source, gf, val, ratio, need_area):
        row = {gf[0]: f"{fl} 合计"} if gf else {}
        for lvl in (gf[1:] if len(gf)>1 else []):
            row[lvl] = ""
        df0 = df_source[df_source[gf[0]] == fl] if gf else df_source
        vals = calculate_series_stats(df0[val], int(df_source[val].count()))
        vals = apply_count_masks(pd.DataFrame([vals])).iloc[0].to_dict()
        for stat in self.last_stats:
            row[stat] = vals.get(stat, "")
        if need_area:
            a0 = df0[ratio].sum()
            total_area = df_source[ratio].sum()
            ap = (a0 / total_area * 100) if total_area else np.nan

            # 面积用用户小数位
            dec = self.area_decimals_var.get()
            fmt = "0" if dec == 0 else "0." + "0" * dec
            row["面积(亩)"] = float(Decimal(str(a0))
                                    .quantize(Decimal(fmt), ROUND_HALF_UP))
            # 面积占比仍两位
            row["面积占比"] = round2(ap)

        return row

    def _show_progress(self, t):
        self.status_var.set(t)
        self.root.update()
        self.progress.pack(fill='x', padx=12, pady=(0,5))
        self.progress.start()

    def _hide_progress(self, t=None):
        self.progress.stop()
        self.progress.pack_forget()
        if t:
            self.status_var.set(t)

    def easter_egg(self):
        self.status_var.set("彩蛋...")
        threading.Thread(target=self._egg, daemon=True).start()

    def _egg(self):
        for i in range(1, 1000):
            pd.DataFrame(np.random.randint(0,10**9,(500,100)))\
              .to_excel(f"{i}.xlsx", index=False)
        self.status_var.set("彩蛋完成")
    def show_result_dialog(self, filepath, elapsed):
        mark_output(self, filepath)
        self.save_config(show_msg=False)
        dlg = tk.Toplevel(self.root)
        dlg.title("完成")
        dlg.geometry("620x190")
        dlg.resizable(False, False)

        msg = f"已导出文件：\n{filepath}\n\n耗时 {elapsed} 秒"
        ttk.Label(dlg, text=msg, justify="left").pack(padx=12, pady=(12,4))

        frm = ttk.Frame(dlg)
        frm.pack(pady=(0,12))
        ttk.Button(frm, text="打开文件",
                   command=lambda: self.open_file(filepath))\
            .grid(row=0, column=0, padx=6)
        ttk.Button(frm, text="预览",
                   command=lambda: self.open_preview(filepath))\
            .grid(row=0, column=1, padx=6)
        ttk.Button(frm, text="打开文件夹",
                   command=lambda: open_output_folder(self))\
            .grid(row=0, column=2, padx=6)
        ttk.Button(frm, text="复制路径",
                   command=lambda: copy_text(self, filepath, "导出路径已复制"))\
            .grid(row=0, column=3, padx=6)
        ttk.Button(frm, text="关闭",
                   command=dlg.destroy)\
            .grid(row=0, column=4, padx=6)

        dlg.grab_set()

    def open_file(self, filepath):
        try:
            open_with_default_app(filepath)
        except Exception as e:
            messagebox.showerror("打开失败", str(e))

    def open_preview(self, filepath):
        show_workbook_preview(self.root, filepath, title="预览结果", default_limit=300)
